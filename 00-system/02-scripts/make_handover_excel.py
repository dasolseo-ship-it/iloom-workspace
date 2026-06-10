import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

OUTPUT = r"c:\Users\FURSYS\Downloads\iloom-workspace-claude\20-operations\2026-06-10_인수인계서_현대목동점_롯데영등포점.xlsx"

wb = openpyxl.Workbook()

# ── 공통 스타일 ────────────────────────────────────────────────
BRAND_RED   = "C80A1E"
DARK_GRAY   = "333333"
MID_GRAY    = "666666"
LIGHT_GRAY  = "F2F2F2"
WHITE       = "FFFFFF"
YELLOW_WARN = "FFF2CC"
SECTION_BG  = "F5F5F5"

def side(color="CCCCCC", style="thin"):
    return Side(border_style=style, color=color)

def full_border(color="CCCCCC"):
    s = side(color)
    return Border(left=s, right=s, top=s, bottom=s)

def header_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=10, color=DARK_GRAY, italic=False):
    return Font(name="맑은 고딕", bold=bold, size=size, color=color, italic=italic)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def write_cell(ws, row, col, value, bold=False, size=10, fg=None, font_color=DARK_GRAY,
               h_align="left", wrap=False, border=True, italic=False, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font  = font(bold=bold, size=size, color=font_color, italic=italic)
    cell.alignment = align(h_align, wrap=wrap)
    if fg:
        cell.fill = header_fill(fg)
    if border:
        cell.border = full_border()
    if number_format:
        cell.number_format = number_format
    return cell

def section_title(ws, row, col_start, col_end, title):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=title)
    cell.font      = font(bold=True, size=11, color=WHITE)
    cell.fill      = header_fill(BRAND_RED)
    cell.alignment = align("left")
    cell.border    = full_border(BRAND_RED)

def sub_header(ws, row, col_start, col_end, title):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=title)
    cell.font      = font(bold=True, size=10, color=DARK_GRAY)
    cell.fill      = header_fill(SECTION_BG)
    cell.alignment = align("left")
    cell.border    = full_border()

def table_header(ws, row, cols_labels, fg=LIGHT_GRAY):
    for col_idx, label in enumerate(cols_labels, start=1):
        write_cell(ws, row, col_idx, label, bold=True, fg=fg,
                   font_color=DARK_GRAY, h_align="center")

# ══════════════════════════════════════════════════════════════
#  SHEET 1 — 현대목동점
# ══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "현대목동점"
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 20
ws1.column_dimensions["B"].width = 38
ws1.column_dimensions["C"].width = 20
ws1.column_dimensions["D"].width = 38

r = 1

# ── 문서 제목 ──
ws1.merge_cells(f"A{r}:D{r}")
c = ws1.cell(row=r, column=1, value="인수인계서 — 일룸 현대목동점")
c.font      = font(bold=True, size=14, color=WHITE)
c.fill      = header_fill(BRAND_RED)
c.alignment = align("center")
c.border    = full_border(BRAND_RED)
ws1.row_dimensions[r].height = 30
r += 1

# 메타
meta = [
    ("작성일", "2026-06-10"),
    ("인계자", "서다솔 (일룸사업부 리테일사업팀, 대리점파트)"),
    ("사유", "담당 매장 변경 (2026-06-04 업무분장 기준)"),
]
for k, v in meta:
    ws1.merge_cells(f"C{r}:D{r}")
    write_cell(ws1, r, 1, k, bold=True, fg=LIGHT_GRAY, h_align="center")
    ws1.merge_cells(f"B{r}:D{r}")
    write_cell(ws1, r, 2, v)
    r += 1
r += 1  # 빈 행

# ── 1. 매장 기본 정보 ──
section_title(ws1, r, 1, 4, "1. 매장 기본 정보")
r += 1
basic = [
    ("매장명",   "일룸 현대목동점",           "매장유형",  "투자형 B삽 (입점)"),
    ("백화점",   "한무쇼핑(주) 목동점 (현대백화점 목동점)", "오픈일", "2026년 5월 1일"),
    ("주소",     "서울특별시 양천구 목동동로 257 현대백화점 목동점 B1층", "면적", "87.27㎡ (약 26.4평)"),
    ("운영층",   "본관 지하1층 (B1F) 리빙존", "",           ""),
]
for k1, v1, k2, v2 in basic:
    write_cell(ws1, r, 1, k1, bold=True, fg=LIGHT_GRAY, h_align="center")
    write_cell(ws1, r, 2, v1)
    write_cell(ws1, r, 3, k2, bold=True, fg=LIGHT_GRAY, h_align="center")
    write_cell(ws1, r, 4, v2)
    r += 1
r += 1

# ── 2. 주요 연락처 ──
section_title(ws1, r, 1, 4, "2. 주요 연락처")
r += 1

# 대리점
sub_header(ws1, r, 1, 4, "▶ 대리점")
r += 1
table_header(ws1, r, ["이름", "역할", "연락처", "비고"])
r += 1
contacts_store = [
    ("안중우", "대표 (사장님)", "—", "사업자번호: 417-10-72598\n유진님 휴무 시 및 수시 방문·근무"),
    ("방현호", "메인 매니저", "010-4757-5877", "소고부터 3년 이상 함께한 매니저\n현호님 휴무 시 → 김정은 사장님 대체"),
    ("유진",   "서브 매니저", "—", "부천점 → 현대중동 팝업 → 현재 현대목동 근무"),
    ("김정은", "사장님 (방현호 휴무 대체)", "—", "방현호 매니저 휴무 시 매장 근무"),
]
for row_data in contacts_store:
    for col_idx, val in enumerate(row_data, start=1):
        write_cell(ws1, r, col_idx, val, wrap=True)
    ws1.row_dimensions[r].height = 30
    r += 1
r += 1

# 현대백화점 담당자
sub_header(ws1, r, 1, 4, "▶ 현대백화점 담당자")
r += 1
table_header(ws1, r, ["이름", "역할", "연락처", "이메일"])
r += 1
contacts_dept = [
    ("윤용철 선임", "리빙파트 담당", "010-9485-8603", "yunyyy0810@thehyundai.com"),
    ("정지영",      "특약매입 계약 담당", "—", "H-Partners 전자계약 담당"),
]
for row_data in contacts_dept:
    for col_idx, val in enumerate(row_data, start=1):
        write_cell(ws1, r, col_idx, val)
    r += 1
r += 1

# ── 3. 계약 현황 ──
section_title(ws1, r, 1, 4, "3. 계약 현황")
r += 1

sub_header(ws1, r, 1, 4, "▶ 위탁운영 계약 (품의번호: 일룸-품의26-04-00113)")
r += 1
contract1 = [
    ("계약기간", "2026.05.01 ~ 2027.03.31 (11개월)", "거래보증금", "20,000,000원"),
    ("판매수수료율", "A구간 11% / B구간 8% / C구간 6% / D구간 5%", "인테리어", "82,000,000원 (본사 100%, 소울디앤아이)"),
]
for k1, v1, k2, v2 in contract1:
    write_cell(ws1, r, 1, k1, bold=True, fg=LIGHT_GRAY, h_align="center")
    write_cell(ws1, r, 2, v1, wrap=True)
    write_cell(ws1, r, 3, k2, bold=True, fg=LIGHT_GRAY, h_align="center")
    write_cell(ws1, r, 4, v2, wrap=True)
    ws1.row_dimensions[r].height = 24
    r += 1
r += 1

sub_header(ws1, r, 1, 4, "▶ 현대백화점 특약매입 계약 (H-Partners 전자계약)")
r += 1
contract2 = [
    ("계약주체", "한무쇼핑(주)목동점 ↔ (주)일룸", "계약형태", "특약매입 단기거래"),
    ("1차 계약기간", "2026.05.01 ~ 2026.07.31 ★만료 임박", "2차 계약기간", "2026.08.01 ~ 2027.07.31"),
    ("마진율", "정상 15%", "정산기준", "월 판매마감일로부터 40일 이내"),
    ("대금 입금계좌", "신한은행 100022715251 ((주)일룸)", "", ""),
]
for k1, v1, k2, v2 in contract2:
    write_cell(ws1, r, 1, k1, bold=True, fg=LIGHT_GRAY, h_align="center")
    write_cell(ws1, r, 2, v1, wrap=True)
    write_cell(ws1, r, 3, k2, bold=True, fg=LIGHT_GRAY, h_align="center")
    write_cell(ws1, r, 4, v2)
    r += 1

# 경고 셀
ws1.merge_cells(f"A{r}:D{r}")
warn_cell = ws1.cell(row=r, column=1,
    value="⚠ 1차 특약매입 계약 2026-07-31 만료 → 2차 계약 전환 처리 필요 (H-Partners 확인)")
warn_cell.font      = font(bold=True, color="7F4700")
warn_cell.fill      = header_fill(YELLOW_WARN)
warn_cell.alignment = align("center")
warn_cell.border    = full_border("E0A000")
r += 1
r += 1

# ── 4. 보험 현황 ──
section_title(ws1, r, 1, 4, "4. 보험 등록 현황 (재산종합보험)")
r += 1
insurance = [
    ("등록 완료일", "2026-05-15", "등록 특약", "시설소유배상 / 임차자배상 / 구내치료비 (3종)"),
    ("등록 주소", "서울특별시 양천구 목동동로 257 현대백화점 목동점 B1층", "면적", "87.27㎡"),
    ("건물구조", "외벽: 유리 / 기둥: 철근콘크리트 / 지붕: 콘크리트", "", ""),
]
for k1, v1, k2, v2 in insurance:
    write_cell(ws1, r, 1, k1, bold=True, fg=LIGHT_GRAY, h_align="center")
    write_cell(ws1, r, 2, v1, wrap=True)
    write_cell(ws1, r, 3, k2, bold=True, fg=LIGHT_GRAY, h_align="center")
    write_cell(ws1, r, 4, v2)
    ws1.row_dimensions[r].height = 24
    r += 1
ws1.merge_cells(f"A{r}:D{r}")
note = ws1.cell(row=r, column=1, value="※ 매년 보험 갱신 시 특약 3종 모두 재등록 필수")
note.font      = font(italic=True, color=MID_GRAY)
note.fill      = header_fill(YELLOW_WARN)
note.alignment = align("left")
note.border    = full_border("E0A000")
r += 1
r += 1

# ── 5. 임직원 할인 이벤트 ──
section_title(ws1, r, 1, 4, "5. 임직원 할인 이벤트")
r += 1
event_info = [
    ("관리 시트", "https://docs.google.com/spreadsheets/d/1WAYtJYRfTGKWAeLbf8D1EgZkgYNTK8iqfsKJ5BHtYC4/edit?gid=909029606#gid=909029606"),
    ("후속 처리", "행사 전체 종료 후 단가정정 협조전 한 번에 작성하여 기록 남길 것"),
]
for k, v in event_info:
    write_cell(ws1, r, 1, k, bold=True, fg=LIGHT_GRAY, h_align="center")
    ws1.merge_cells(f"B{r}:D{r}")
    write_cell(ws1, r, 2, v, wrap=True)
    ws1.row_dimensions[r].height = 24
    r += 1
r += 1

# ── 6. SNS 인증샷 이벤트 ──
section_title(ws1, r, 1, 4, "6. SNS 인증샷 이벤트 (럭키드로우)")
r += 1
sns_detail = [
    (
        "1·2등 경품\n(현물 관리)",
        "1·2등 당첨 상품이 현물이라 별도 관리 필요.\n→ 별도 시트에 아카이빙 진행 중 (수령 여부·재고 현황 기록 유지할 것)"
    ),
    (
        "캡처본 전달\n방침",
        "담당자 개인 SNS 계정이 캡처에 노출될 수 있음.\n"
        "또한 캡처 + 전달까지 요구하면 고객이 이벤트를 번거롭게 느껴 피로감 유발 가능.\n"
        "→ 캡처본에 계정 노출은 필수 아님\n"
        "→ '캡처 싫다'고 거부하는 고객은 강요 없이 이벤트 생략 가능"
    ),
    (
        "사전 허락\n완료",
        "위 방침은 사전에 권희님·재경팀에 공유 및 허락 완료.\n→ 신규 담당자도 동일 방침 적용 가능 (별도 재확인 불필요)"
    ),
    (
        "운영 목적",
        "캡처 수집 자체가 목적이 아님.\n"
        "\"본사에서 모니터링하고 있으니 허투루 쓰지 말라\"는 긴장감 유지 용도.\n"
        "→ 굳이 캡처를 독려하거나 강제할 필요 없이, 자연스럽게 지켜보는 정도로 운영"
    ),
]
for k, v in sns_detail:
    write_cell(ws1, r, 1, k, bold=True, fg=LIGHT_GRAY, h_align="center", wrap=True)
    ws1.merge_cells(f"B{r}:D{r}")
    write_cell(ws1, r, 2, v, wrap=True)
    ws1.row_dimensions[r].height = 60
    r += 1
r += 1

# ── 7. 오픈 정산 현황 ──
section_title(ws1, r, 1, 4, "7. 오픈 정산 현황 (완료)")
r += 1
table_header(ws1, r, ["항목", "본사 부담 (원)", "대리점 부담 (원)", "합계 (원)"])
r += 1
settlement = [
    ("인테리어 (소울디앤아이)",    82000000, 0,       82000000),
    ("내/외부사인 (비웍스디자인)", 2160000,  930000,  3090000),
    ("소품 (브레인크리에이티브)",  1784000,  1784000, 3568000),
    ("포스터 (인성커뮤니케이션)",  36500,    0,       36500),
    ("SNS 인증샷 이벤트",         1496545,  0,       1496545),
    ("구매 이벤트 (가족필통)",    783600,   0,       783600),
]
num_fmt = '#,##0"원"'
for name, company, agency, total in settlement:
    write_cell(ws1, r, 1, name)
    write_cell(ws1, r, 2, company, h_align="right", number_format=num_fmt)
    write_cell(ws1, r, 3, agency,  h_align="right", number_format=num_fmt)
    write_cell(ws1, r, 4, total,   h_align="right", number_format=num_fmt)
    r += 1
# 합계
write_cell(ws1, r, 1, "합계", bold=True, fg=LIGHT_GRAY, h_align="center")
write_cell(ws1, r, 2, 88260645, bold=True, fg=LIGHT_GRAY, h_align="right", number_format=num_fmt)
write_cell(ws1, r, 3, 2714000,  bold=True, fg=LIGHT_GRAY, h_align="right", number_format=num_fmt)
write_cell(ws1, r, 4, 90974645, bold=True, fg=LIGHT_GRAY, h_align="right", number_format=num_fmt)
r += 1
r += 1

# ── 8. 인수 시 체크리스트 ──
section_title(ws1, r, 1, 4, "8. 인수 시 체크리스트")
r += 1
checklist = [
    "1차 특약매입 계약 만료(2026-07-31) → 2차 계약 전환 처리 (H-Partners)",
    "임직원 할인 이벤트 종료 후 단가정정 협조전 작성",
    "매년 보험 갱신 시 특약 3종(시설소유배상·임차자배상·구내치료비) 재등록",
    "위탁운영 계약 만료(2027-03-31) 2~3개월 전 갱신 여부 협의 시작",
]
for item in checklist:
    ws1.merge_cells(f"A{r}:D{r}")
    cell = ws1.cell(row=r, column=1, value=f"☐  {item}")
    cell.font      = font(size=10)
    cell.alignment = align("left", wrap=True)
    cell.fill      = header_fill(YELLOW_WARN)
    cell.border    = full_border("E0A000")
    ws1.row_dimensions[r].height = 24
    r += 1
r += 1

# 품의 링크
sub_header(ws1, r, 1, 4, "▶ 관련 품의 링크")
r += 1
links = [
    ("개설 품의",      "https://ep.fursys.com/WebFlow/view.do?flowNo=ku14LRL7nuP9Vnoy3IteioEs6ZqRxlNqaxLYH7fH3vM"),
    ("오픈 마케팅 품의", "https://ep.fursys.com/WebFlow/view.do?flowNo=LsQBuI0ZOuOmU7QxkRdNYqsqf8PTQmAOEUcqWwbAN84"),
]
for k, v in links:
    write_cell(ws1, r, 1, k, bold=True, fg=LIGHT_GRAY, h_align="center")
    ws1.merge_cells(f"B{r}:D{r}")
    c = ws1.cell(row=r, column=2, value=v)
    c.font = Font(name="맑은 고딕", size=9, color="0563C1", underline="single")
    c.alignment = align("left")
    c.border = full_border()
    r += 1


# ══════════════════════════════════════════════════════════════
#  SHEET 2 — 롯데영등포점
# ══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("롯데영등포점")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 38
ws2.column_dimensions["C"].width = 20
ws2.column_dimensions["D"].width = 38

r = 1

# 제목
ws2.merge_cells(f"A{r}:D{r}")
c = ws2.cell(row=r, column=1, value="인수인계서 — 일룸 롯데영등포점")
c.font      = font(bold=True, size=14, color=WHITE)
c.fill      = header_fill(BRAND_RED)
c.alignment = align("center")
c.border    = full_border(BRAND_RED)
ws2.row_dimensions[r].height = 30
r += 1

meta2 = [
    ("작성일", "2026-06-10"),
    ("인계자", "서다솔 (일룸사업부 리테일사업팀, 대리점파트)"),
    ("사유",   "담당 매장 변경 (2026-06-04 업무분장 기준)"),
]
for k, v in meta2:
    write_cell(ws2, r, 1, k, bold=True, fg=LIGHT_GRAY, h_align="center")
    ws2.merge_cells(f"B{r}:D{r}")
    write_cell(ws2, r, 2, v)
    r += 1
r += 1

# ── 1. 매장 기본 정보 ──
section_title(ws2, r, 1, 4, "1. 매장 기본 정보")
r += 1
basic2 = [
    ("매장명",    "일룸 롯데영등포점",       "매장유형",  "입점BS"),
    ("백화점",    "롯데백화점 영등포점",      "면적",      "53평"),
    ("주소",      "서울특별시 영등포구 영등포동 경인로 846 롯데백화점 영등포점 9층",
     "매장코드",  "62LM26"),
    ("대표전화",  "02-2630-6933",            "",          ""),
]
for k1, v1, k2, v2 in basic2:
    write_cell(ws2, r, 1, k1, bold=True, fg=LIGHT_GRAY, h_align="center")
    write_cell(ws2, r, 2, v1, wrap=True)
    write_cell(ws2, r, 3, k2, bold=True, fg=LIGHT_GRAY, h_align="center")
    write_cell(ws2, r, 4, v2)
    ws2.row_dimensions[r].height = 24
    r += 1
r += 1

# ── 2. 주요 연락처 ──
section_title(ws2, r, 1, 4, "2. 주요 연락처")
r += 1
table_header(ws2, r, ["이름", "역할", "연락처", "비고"])
r += 1
contacts2 = [
    ("오승균", "대표 (사장님) — 메인 소통 창구", "010-5055-0253 / gyun0502@hotmail.com",
     "김지원 매니저 휴무 시 단독 근무"),
    ("김지원", "매니저", "—", "평상시 오승균 사장님과 2인 근무"),
]
for row_data in contacts2:
    for col_idx, val in enumerate(row_data, start=1):
        write_cell(ws2, r, col_idx, val, wrap=True)
    ws2.row_dimensions[r].height = 28
    r += 1
r += 1

# ── 3. B2B 업무 이력 ──
section_title(ws2, r, 1, 4, "3. B2B 업무 이력")
r += 1
b2b_info = [
    ("처리일",    "2026-05-26",          "거래처",  "구세군작업장"),
    ("할인율",    "21%",                  "처리 문서", "협조전 + B2B거래물품할인공급요청서"),
    ("파일 위치", "20-operations/22-B2B업무/\n협조전_롯데영등포_B2B_구세군작업장.docx\nB2B거래물품할인공급요청서_롯데영등포_수정본.docx",
     "", ""),
]
for k1, v1, k2, v2 in b2b_info:
    write_cell(ws2, r, 1, k1, bold=True, fg=LIGHT_GRAY, h_align="center")
    write_cell(ws2, r, 2, v1, wrap=True)
    write_cell(ws2, r, 3, k2, bold=True, fg=LIGHT_GRAY, h_align="center")
    write_cell(ws2, r, 4, v2, wrap=True)
    ws2.row_dimensions[r].height = 40
    r += 1
r += 1

# ── 4. 인수 시 체크리스트 ──
section_title(ws2, r, 1, 4, "4. 인수 시 체크리스트")
r += 1
checklist2 = [
    "구세군작업장 B2B 건 정산 완료 여부 → 오승균 사장님(010-5055-0253)에게 확인",
    "현재 계약 만료일 및 갱신 여부 확인 (계약정서 원본 검토)",
    "진행 중인 B2B 건 추가 여부 확인",
    "5월 판촉비 집행 현황 (포함 6개점 기준 17,056,440원) 이월 항목 확인",
]
for item in checklist2:
    ws2.merge_cells(f"A{r}:D{r}")
    cell = ws2.cell(row=r, column=1, value=f"☐  {item}")
    cell.font      = font(size=10)
    cell.alignment = align("left", wrap=True)
    cell.fill      = header_fill(YELLOW_WARN)
    cell.border    = full_border("E0A000")
    ws2.row_dimensions[r].height = 24
    r += 1

# ══════════════════════════════════════════════════════════════
#  저장
# ══════════════════════════════════════════════════════════════
wb.save(OUTPUT)
print(f"저장 완료: {OUTPUT}")
