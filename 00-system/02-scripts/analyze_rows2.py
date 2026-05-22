import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(
    r'c:\Users\FURSYS\Downloads\iloom-workspace-claude\10-projects\14-hyundaimokdong-marketing\현대목동점_위탁판매_대리점_계약정서_2026.xlsx',
    data_only=False
)

# 약정서(CI,SI) 전체 출력
for name in ['약정서(CI,SI)', '약정서(수수료-투자b입점)']:
    ws = wb[name]
    breaks = [b.id for b in ws.row_breaks.brk] if ws.row_breaks else []
    print(f'\n[{name}] BREAK at {breaks}  max={ws.max_row}')
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            v = str(cell.value).strip() if cell.value else ''
            if v and len(v) > 2:
                marker = ' <<=BREAK' if cell.row in breaks else ''
                print(f'  {cell.row:4d}: {v[:70]}{marker}')
                break
