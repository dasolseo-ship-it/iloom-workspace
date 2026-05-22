"""
페이지 나누기 진단 v2 - 정확한 행 높이 기반 계산
defaultRowHeight=17.0pt 기준
"""
import openpyxl, sys, re
sys.stdout.reconfigure(encoding='utf-8')

PATH = r'10-projects/14-hyundaimokdong-marketing/현대목동점_위탁판매_대리점_계약정서_2026.xlsx'
wb = openpyxl.load_workbook(PATH, data_only=False)

# A4 세로, 여백: 상하 0.71" = 51.12pt, 헤더/푸터 0.3" = 21.6pt
# A4 = 841.89pt 높이
USABLE_HEIGHT_PT = 841.89 - 2*51.12 - 2*21.6  # ≈ 697pt

def get_row_height(ws, row_num):
    rd = ws.row_dimensions.get(row_num)
    if rd and rd.height:
        return rd.height
    dh = ws.sheet_format.defaultRowHeight
    return dh if dh else 17.0

def analyze_sheet_pages(ws, sheet_name, start_page=1):
    breaks = sorted([b.id for b in ws.row_breaks.brk]) if ws.row_breaks else []

    pa = ws.print_area
    if pa:
        m = re.search(r'\$A\$1:\$[A-Z]+\$(\d+)', pa)
        max_print_row = int(m.group(1)) if m else ws.max_row
    else:
        max_print_row = ws.max_row

    print(f'\n{"="*60}')
    print(f'[{sheet_name}] breaks={breaks} print_rows=1~{max_print_row}  (전체페이지시작={start_page})')
    print(f'  usable height={USABLE_HEIGHT_PT:.1f}pt')
    print(f'{"="*60}')

    current_height = 0
    page_num = start_page
    page_row_start = 1
    pages = []

    for row in range(1, max_print_row + 1):
        h = get_row_height(ws, row)

        # 수동 페이지 나누기 (break는 해당 행 이전에서 나눔)
        if row in breaks:
            pages.append((page_num, page_row_start, row - 1))
            print(f'  페이지{page_num}: rows {page_row_start}~{row-1} [수동 break before {row}]')
            page_num += 1
            page_row_start = row
            current_height = 0

        current_height += h

        # 자동 페이지 넘침
        if current_height > USABLE_HEIGHT_PT:
            # 이전 행까지가 이 페이지
            pages.append((page_num, page_row_start, row - 1))
            print(f'  페이지{page_num}: rows {page_row_start}~{row-1} [자동 넘침 at row {row}]')
            page_num += 1
            page_row_start = row
            current_height = h

    # 마지막 페이지
    pages.append((page_num, page_row_start, max_print_row))
    print(f'  페이지{page_num}: rows {page_row_start}~{max_print_row} [마지막]')
    page_num += 1

    print(f'  → 이 시트: {page_num - start_page}페이지')
    return page_num  # 다음 시트 시작 페이지

targets = ['계약서', '약정서(수수료-투자b입점)', '약정서(CI,SI)', '약정서(개인정보)', '3D홈플래너 사용권계약서', '동반성장 및 청렴서약서']

next_page = 1
for name in targets:
    ws = wb[name]
    next_page = analyze_sheet_pages(ws, name, next_page)

print(f'\n총 합계: {next_page - 1}페이지')
