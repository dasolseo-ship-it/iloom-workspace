"""
정확한 페이지 계산 - 실제 행 높이 누적
각 시트별 페이지 경계를 정확히 파악
"""
import openpyxl, sys, re
sys.stdout.reconfigure(encoding='utf-8')

PATH = r'10-projects/14-hyundaimokdong-marketing/현대목동점_위탁판매_대리점_계약정서_2026.xlsx'
wb = openpyxl.load_workbook(PATH, data_only=False)

# A4 세로, 여백: 상하 0.71"=51.12pt, 헤더/푸터 0.3"=21.6pt
# A4 높이 = 841.89pt
USABLE_HEIGHT = 841.89 - 2*51.12 - 2*21.6  # ≈ 696.45pt

def get_h(ws, row_num):
    rd = ws.row_dimensions.get(row_num)
    if rd and rd.height:
        return rd.height
    dh = ws.sheet_format.defaultRowHeight
    return dh if dh else 17.0

def get_first_content(ws, row_num):
    for cell in ws[row_num]:
        v = str(cell.value).strip() if cell.value else ''
        if v:
            return v[:60]
    return ''

def analyze(ws, name, global_start=1):
    breaks = sorted([b.id for b in ws.row_breaks.brk]) if ws.row_breaks else []
    pa = ws.print_area
    if pa:
        m = re.search(r'\$A\$1:\$[A-Z]+\$(\d+)', pa)
        max_row = int(m.group(1)) if m else ws.max_row
    else:
        max_row = ws.max_row

    print(f'\n{"="*60}')
    print(f'[{name}] breaks={breaks} rows=1~{max_row} start_page={global_start}')
    print(f'{"="*60}')

    pages = []
    cum_h = 0.0
    page_start_row = 1
    pg = global_start

    for r in range(1, max_row + 1):
        h = get_h(ws, r)

        # 수동 break: 이 행의 바로 앞에서 페이지 나누기
        if r in breaks:
            content_prev = get_first_content(ws, r - 1) if r > 1 else ''
            content_here = get_first_content(ws, r)
            print(f'  📄페이지{pg}: rows {page_start_row}~{r-1} (마지막내용: {content_prev[:40]})')
            print(f'     → break at {r}: {content_here[:40]}')
            pages.append((pg, page_start_row, r-1))
            pg += 1
            page_start_row = r
            cum_h = 0.0

        cum_h += h

        # 자동 overflow
        if cum_h > USABLE_HEIGHT:
            prev_content = get_first_content(ws, r - 1) if r > 1 else ''
            curr_content = get_first_content(ws, r)
            print(f'  📄페이지{pg}: rows {page_start_row}~{r-1} (auto overflow, 누적={cum_h:.0f}pt, 마지막: {prev_content[:30]})')
            pages.append((pg, page_start_row, r-1))
            pg += 1
            page_start_row = r
            cum_h = h

    # 마지막 페이지
    last_content = get_first_content(ws, max_row)
    print(f'  📄페이지{pg}: rows {page_start_row}~{max_row} [마지막] ({last_content[:40]})')
    pages.append((pg, page_start_row, max_row))
    pg += 1

    print(f'  → {name}: {pg - global_start}페이지 (p{global_start}~p{pg-1})')
    return pg

targets = ['계약서', '약정서(수수료-투자b입점)', '약정서(CI,SI)', '약정서(개인정보)', '3D홈플래너 사용권계약서', '동반성장 및 청렴서약서']
nxt = 1
for n in targets:
    nxt = analyze(wb[n], n, nxt)

print(f'\n총 {nxt-1}페이지')
