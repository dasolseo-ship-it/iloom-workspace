import openpyxl

PATH = r'c:\Users\FURSYS\Downloads\iloom-workspace-claude\10-projects\14-hyundaimokdong-marketing\현대목동점_위탁판매_대리점_계약정서_2026.xlsx'

wb = openpyxl.load_workbook(PATH, data_only=False)

OLD = '경기도 고양시 덕양구 고양대로 1955, B1층(동산동)'
NEW = '서울 양천구 목동동로 257 현대백화점 목동점 지하1층'

count = 0
for sheet in wb.sheetnames:
    ws = wb[sheet]
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and OLD in str(cell.value):
                cell.value = str(cell.value).replace(OLD, NEW)
                print(f'  수정 [{sheet}] {cell.coordinate}')
                count += 1

wb.save(PATH)
print(f'\n총 {count}개 셀 수정 완료 ✓')
