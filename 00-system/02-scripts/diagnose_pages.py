"""
페이지 나누기 진단 - 전체 페이지 구조 파악
각 시트별 행 높이를 기준으로 페이지 분포 계산
"""
import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

PATH = r'10-projects/14-hyundaimokdong-marketing/현대목동점_위탁판매_대리점_계약정서_2026.xlsx'
wb = openpyxl.load_workbook(PATH, data_only=False)

# A4 세로, 여백 상하 0.71" (51pt), 헤더/푸터 0.3" (21.6pt)
# A4 높이 = 841pt, 실제 인쇄 가능 높이 = 841 - 2*51 - 2*21.6 ≈ 696pt
# 기본 행 높이 = 14.4pt
USABLE_HEIGHT_PT = 696  # 사용 가능 높이 (포인트)
DEFAULT_ROW_HEIGHT = 14.4  # 기본 행 높이 (포인트)

def get_row_height(ws, row_num):
    rd = ws.row_dimensions.get(row_num)
    if rd and rd.height:
        return rd.height
    return DEFAULT_ROW_HEIGHT

def analyze_sheet_pages(ws, sheet_name):
    breaks = sorted([b.id for b in ws.row_breaks.brk]) if ws.row_breaks else []

    # print_area에서 max_row 파악
    pa = ws.print_area
    if pa:
        import re
        m = re.search(r'\$A\$1:\$[A-Z]+\$(\d+)', pa)
        max_print_row = int(m.group(1)) if m else ws.max_row
    else:
        max_print_row = ws.max_row

    print(f'\n{"="*60}')
    print(f'[{sheet_name}] breaks={breaks} print_rows=1~{max_print_row}')
    print(f'{"="*60}')

    current_height = 0
    page_num = 1
    page_starts = [1]

    for row in range(1, max_print_row + 1):
        h = get_row_height(ws, row)

        # 수동 페이지 나누기 확인
        if row in breaks and current_height > 0:
            print(f'  페이지{page_num}: rows 1~{row-1} (break at {row})')
            page_num += 1
            page_starts.append(row)
            current_height = h
            continue

        current_height += h

        # 자동 페이지 넘침
        if current_height > USABLE_HEIGHT_PT:
            print(f'  페이지{page_num}: rows ~{row-1} (auto overflow, height={current_height:.0f}pt)')
            page_num += 1
            page_starts.append(row)
            current_height = h

    print(f'  페이지{page_num}: rows ~{max_print_row} (마지막)')
    print(f'  → 총 {page_num}페이지')
    return page_num

targets = ['계약서', '약정서(수수료-투자b입점)', '약정서(CI,SI)', '약정서(개인정보)', '3D홈플래너 사용권계약서', '동반성장 및 청렴서약서']

total = 0
for name in targets:
    ws = wb[name]
    n = analyze_sheet_pages(ws, name)
    total += n

print(f'\n총 합계: {total}페이지')
