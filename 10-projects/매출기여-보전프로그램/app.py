"""
매출기여 보전 프로그램
온라인 행사로 인한 오프라인 수주 취소건 보전금 관리
"""
from __future__ import annotations
import os
import re
import json
import psycopg2
import psycopg2.extras
import psycopg2.pool
import io
from threading import Lock
from datetime import date, datetime, timedelta
from typing import List, Optional
from contextlib import contextmanager

import uvicorn
from fastapi import FastAPI, HTTPException, Request, UploadFile, File, Body
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel

from contextlib import asynccontextmanager

@asynccontextmanager
async def lifespan(application):
    try:
        init_db()
    except Exception:
        import traceback
        print("=== DB 초기화 실패 ===")
        print(traceback.format_exc())
        print("=====================")
    yield

app = FastAPI(title="매출기여 보전 프로그램", lifespan=lifespan)

_HERE = os.path.dirname(os.path.abspath(__file__))
DATABASE_URL = os.environ.get("DATABASE_URL", "")
templates = Jinja2Templates(directory=os.path.join(_HERE, "templates"))

_pool: Optional[psycopg2.pool.ThreadedConnectionPool] = None
_pool_lock = Lock()

def _get_pool() -> psycopg2.pool.ThreadedConnectionPool:
    global _pool
    if _pool is not None and not _pool.closed:
        return _pool
    with _pool_lock:
        if _pool is None or _pool.closed:
            _pool = psycopg2.pool.ThreadedConnectionPool(
                1, 4, DATABASE_URL,
                keepalives=1, keepalives_idle=30,
                keepalives_interval=10, keepalives_count=3,
            )
    return _pool

# ================================================================
# 일룸 시리즈 / 품목 설정
# ================================================================
SERIES_LIST = sorted([
    "쿠시노코지", "쿠시노",
    "로이모노", "로이",
    "에디키즈", "에디",
    "바젤SS", "바젤", "반트SS", "반트",
    "멘디", "클러치", "포레스트", "렉스",
    "퍼스트", "파로", "라비", "워드",
    "스탠다", "학생방", "기타",
], key=len, reverse=True)  # 긴 이름 우선 매칭

CATEGORIES = [
    "침대", "소파", "책상", "의자", "수납장", "옷장",
    "서랍장", "화장대", "거울", "선반", "책장", "식탁",
    "파티션", "기타",
]

# 온라인 채널 분류 기준 (태블로 실적대리점 기준 포함)
ONLINE_STORES = frozenset([
    '네이버', '일룸쇼핑몰', '쿠팡', '쿠팡로켓', '엘롯데', 'LG홈스타일',
    'EB', 'AS', '온라인사업부',
    'NC대전유성', 'NC대전유성서브',
    '오늘의 집', '오늘의집', '29CM', '에스에스지', '씨제이몰', '더현대닷컴',
])

# 직영매장 제외 목록 — 태블로 업로드 시 자동 스킵
DIRECT_STORES = frozenset([
    '논현4', '분당서현', '송파4', '용산2', '미포2', '노원',
    '강동아이파크', '부산센텀4', '수원광교3', '대전둔산5', '대구3',
])

# ERP 납기완료 상태값 (시스템마다 표기 다를 수 있어 여러 값 허용)
DELIVERY_DONE_STATUSES = frozenset(['납기완료', '출고완료', '배송완료', '완료', '출고'])

# 담당 오프라인 매장 (CLAUDE.md 기준 — 2026-06-04 업무분장)
MY_STORES = frozenset([
    '인천중앙2', '인천검단', '김포5', '의정부8', '롯데구리', '롯데인천2', '부천3',
    '춘천3', '원주3', '덕진2', '중화산4', '전군2', '송도5',
])

# 쿠시노 ≠ 쿠시노코지, 로이 ≠ 로이모노
EXCLUSION_RULES: dict[str, list[str]] = {
    "쿠시노": ["쿠시노코지"],
    "쿠시노코지": ["쿠시노"],
    "로이": ["로이모노"],
    "로이모노": ["로이"],
}


# ================================================================
# DB
# ================================================================
class _Conn:
    """psycopg2 connection wrapper — sqlite3-like interface for drop-in compatibility."""

    def __init__(self, pg_conn):
        self._pg = pg_conn

    def execute(self, sql, params=()):
        sql = sql.replace('?', '%s')
        cur = self._pg.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(sql, list(params) if params else None)
        return cur

    def executemany(self, sql, seq):
        sql = sql.replace('?', '%s')
        cur = self._pg.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        psycopg2.extras.execute_batch(cur, sql, seq)
        return cur

    def commit(self):
        self._pg.commit()

    def rollback(self):
        self._pg.rollback()

    def close(self):
        self._pg.close()


@contextmanager
def get_db():
    global _pool
    conn, pool, from_pool = None, None, False
    try:
        pool = _get_pool()
        conn = pool.getconn()
        from_pool = True
    except Exception:
        conn = psycopg2.connect(DATABASE_URL)
    wrapped = _Conn(conn)
    try:
        yield wrapped
        conn.commit()
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        raise
    finally:
        if from_pool and pool and not pool.closed:
            try:
                pool.putconn(conn)
            except Exception:
                try: conn.close()
                except Exception: pass
        else:
            try: conn.close()
            except Exception: pass


def init_db():
    with get_db() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS events (
                id SERIAL PRIMARY KEY,
                event_name TEXT NOT NULL,
                announcement_date TEXT NOT NULL,
                start_date TEXT NOT NULL,
                end_date TEXT NOT NULL,
                online_channel TEXT DEFAULT '네이버',
                offline_lookback_days INTEGER DEFAULT 1,
                created_at TEXT DEFAULT TO_CHAR(NOW(), 'YYYY-MM-DD HH24:MI:SS')
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS event_products (
                id SERIAL PRIMARY KEY,
                event_id INTEGER NOT NULL REFERENCES events(id) ON DELETE CASCADE,
                product_name TEXT NOT NULL,
                series TEXT NOT NULL,
                category TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS orders (
                id SERIAL PRIMARY KEY,
                store_name TEXT NOT NULL,
                order_date TEXT NOT NULL,
                customer_amount REAL NOT NULL,
                event_id INTEGER REFERENCES events(id),
                result_type TEXT,
                compensation REAL DEFAULT 0,
                note TEXT,
                created_at TEXT DEFAULT TO_CHAR(NOW(), 'YYYY-MM-DD HH24:MI:SS')
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS order_products (
                id SERIAL PRIMARY KEY,
                order_id INTEGER NOT NULL REFERENCES orders(id) ON DELETE CASCADE,
                product_name TEXT NOT NULL,
                series TEXT NOT NULL,
                category TEXT NOT NULL,
                match_status TEXT DEFAULT 'pending'
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS erp_orders (
                order_no TEXT PRIMARY KEY,
                order_base TEXT NOT NULL,
                order_seq INTEGER NOT NULL DEFAULT 0,
                store_name TEXT NOT NULL,
                store_type TEXT NOT NULL,
                order_status TEXT NOT NULL,
                customer_name TEXT DEFAULT '',
                order_name TEXT DEFAULT '',
                order_date TEXT,
                delivery_date TEXT,
                amount REAL DEFAULT 0,
                address_dong TEXT,
                phone_last4 TEXT,
                cancel_type TEXT,
                import_date TEXT DEFAULT TO_CHAR(CURRENT_DATE, 'YYYY-MM-DD')
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS matches (
                id SERIAL PRIMARY KEY,
                offline_order_no TEXT NOT NULL,
                online_order_no TEXT NOT NULL,
                event_id INTEGER REFERENCES events(id),
                match_keys TEXT,
                match_confidence TEXT DEFAULT 'medium',
                result_type TEXT,
                compensation REAL DEFAULT 0,
                status TEXT DEFAULT 'pending',
                created_at TEXT DEFAULT TO_CHAR(NOW(), 'YYYY-MM-DD HH24:MI:SS')
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS match_items (
                id SERIAL PRIMARY KEY,
                match_id INTEGER NOT NULL REFERENCES matches(id) ON DELETE CASCADE,
                product_name TEXT DEFAULT '',
                series TEXT NOT NULL,
                category TEXT NOT NULL,
                amount REAL DEFAULT 0,
                match_status TEXT DEFAULT 'pending'
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS erp_order_lines (
                id SERIAL PRIMARY KEY,
                order_no TEXT NOT NULL,
                set_code TEXT DEFAULT '',
                series TEXT DEFAULT '',
                category TEXT DEFAULT '',
                sub_product TEXT DEFAULT '',
                sub_product2 TEXT DEFAULT '',
                sub_product3 TEXT DEFAULT '',
                amount REAL DEFAULT 0
            )
        """)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_erp_store_type    ON erp_orders(store_type)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_erp_customer_dong  ON erp_orders(customer_name, address_dong)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_erp_order_date    ON erp_orders(order_date)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_erp_cancel_type   ON erp_orders(cancel_type)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_matches_event     ON matches(event_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_lines_order_no    ON erp_order_lines(order_no)")

        # 기존 DB 마이그레이션 (컬럼 누락 시 추가)
        ev_cols = [r["column_name"] for r in conn.execute("""
            SELECT column_name FROM information_schema.columns
            WHERE table_name='events' AND table_schema='public'
        """).fetchall()]
        if "online_channel" not in ev_cols:
            conn.execute("ALTER TABLE events ADD COLUMN online_channel TEXT DEFAULT '네이버'")
        if "offline_lookback_days" not in ev_cols:
            conn.execute("ALTER TABLE events ADD COLUMN offline_lookback_days INTEGER DEFAULT 7")
        if "is_closed" not in ev_cols:
            conn.execute("ALTER TABLE events ADD COLUMN is_closed BOOLEAN DEFAULT FALSE")
        if "closed_at" not in ev_cols:
            conn.execute("ALTER TABLE events ADD COLUMN closed_at TEXT DEFAULT NULL")
        if "is_deleted" not in ev_cols:
            conn.execute("ALTER TABLE events ADD COLUMN is_deleted BOOLEAN DEFAULT FALSE")
        if "deleted_at" not in ev_cols:
            conn.execute("ALTER TABLE events ADD COLUMN deleted_at TEXT DEFAULT NULL")

        item_cols = [r["column_name"] for r in conn.execute("""
            SELECT column_name FROM information_schema.columns
            WHERE table_name='match_items' AND table_schema='public'
        """).fetchall()]
        if "sub_product" not in item_cols:
            conn.execute("ALTER TABLE match_items ADD COLUMN sub_product TEXT DEFAULT ''")
        if "sub_product2" not in item_cols:
            conn.execute("ALTER TABLE match_items ADD COLUMN sub_product2 TEXT DEFAULT ''")
        if "sub_product3" not in item_cols:
            conn.execute("ALTER TABLE match_items ADD COLUMN sub_product3 TEXT DEFAULT ''")

        conn.execute("""
            CREATE UNIQUE INDEX IF NOT EXISTS idx_matches_pair
            ON matches(offline_order_no, online_order_no)
        """)


# ================================================================
# 태블로 헬퍼 함수
# ================================================================
def parse_korean_date(v) -> str | None:
    """'2026년 4월 15일' → '2026-04-15'"""
    if v is None:
        return None
    if hasattr(v, 'strftime'):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    m = re.match(r'(\d{4})년\s*(\d{1,2})월\s*(\d{1,2})일', s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return s[:10] if len(s) >= 10 else None


def extract_phone_last4_tableau(phone_str) -> str | None:
    """'010-****-7293' → '7293'"""
    if not phone_str:
        return None
    m = re.search(r'-(\d{4})$', str(phone_str).strip())
    return m.group(1) if m else None


def extract_customer_name_tableau(name_str) -> str:
    """'손은정(폐기장무상)(윤)' → '손은정'"""
    if not name_str:
        return ""
    m = re.match(r'^([가-힣]{2,5})', str(name_str).strip())
    return m.group(1) if m else ""


# ================================================================
# ERP 헬퍼 함수
# ================================================================
def extract_customer_name(order_name: str) -> str:
    """수주건명에서 고객명 추출. 예) (온라인택배)이상이(이*름) → 이상이"""
    if not order_name:
        return ""
    s = re.sub(r'^(\([^)]*\)\s*)+', '', order_name.strip())
    m = re.match(r'^([가-힣]{2,4})(?:\(|$)', s)
    return m.group(1) if m else ""


def parse_order_no(order_no: str) -> tuple[str, int]:
    """I20260529-0088-01 → ('I20260529-0088', 1)"""
    m = re.match(r'^(.*)-(\d+)$', str(order_no).strip())
    if m:
        return m.group(1), int(m.group(2))
    return order_no, 0


def classify_cancel_types(conn):
    """취소건을 date_change / pure_cancel 로 분류"""
    all_rows = conn.execute(
        "SELECT order_no, order_base, order_seq FROM erp_orders"
    ).fetchall()

    base_max: dict[str, int] = {}
    for r in all_rows:
        b = r["order_base"]
        if b not in base_max or r["order_seq"] > base_max[b]:
            base_max[b] = r["order_seq"]

    for r in conn.execute(
        "SELECT order_no, order_base, order_seq FROM erp_orders WHERE order_status='취소'"
    ).fetchall():
        ct = "date_change" if r["order_seq"] < base_max.get(r["order_base"], 0) else "pure_cancel"
        conn.execute("UPDATE erp_orders SET cancel_type=? WHERE order_no=?", (ct, r["order_no"]))


def run_matching_engine(conn) -> dict:
    """
    단일 JOIN으로 후보 추출 → 배치 INSERT (N+1 제거)
    result_type:
      cancel_match_delivered : 오프라인 취소 + 온라인 납기완료 → 보전 확정
      cancel_match_pending   : 오프라인 취소 + 온라인 수주중   → 보전 대기
      active_match           : 오프라인 수주중 + 온라인 수주   → 참고
    """
    events = conn.execute("SELECT * FROM events ORDER BY announcement_date").fetchall()
    total_matched = 0
    today_str = date.today().isoformat()

    for ev in events:
        ann_date = date.fromisoformat(ev["announcement_date"])
        end_date = date.fromisoformat(ev["end_date"])
        offline_from = (ann_date - timedelta(days=1)).isoformat()

        cutoff_m = end_date.month + 1
        cutoff_y = end_date.year + (cutoff_m - 1) // 12
        cutoff_m = ((cutoff_m - 1) % 12) + 1
        delivery_cutoff = end_date.replace(year=cutoff_y, month=cutoff_m).isoformat()

        # 단일 JOIN: customer_name + address_dong 기준 후보 쌍 전부 추출
        # online_channel 조건: 행사가 진행된 채널(플랫폼)에 한정하여 인정
        candidates = conn.execute("""
            SELECT
                off.order_no    AS off_no,
                off.order_status AS off_status,
                off.cancel_type,
                off.amount      AS off_amount,
                off.phone_last4  AS off_phone,
                on_.order_no    AS on_no,
                on_.order_status AS on_status,
                on_.delivery_date AS on_delivery,
                on_.phone_last4  AS on_phone
            FROM erp_orders on_
            JOIN erp_orders off
              ON  off.customer_name = on_.customer_name
              AND off.address_dong  = on_.address_dong
              AND off.store_type    = 'offline'
              AND off.order_date   >= ?
              AND off.order_date   <= ?
              AND off.order_date   <= on_.order_date
            WHERE on_.store_type   = 'online'
              AND on_.store_name   = ?
              AND on_.order_date   >= ?
              AND on_.order_date   <= ?
              AND on_.customer_name != ''
              AND on_.address_dong  IS NOT NULL
              AND on_.order_status  != '취소'
              AND (on_.delivery_date IS NULL OR on_.delivery_date <= ?)
        """, (offline_from, ev["end_date"],
              ev["online_channel"],
              ev["start_date"], ev["end_date"],
              delivery_cutoff)).fetchall()

        # 중복 제거: 이미 등록된 매칭 키 세트
        existing = set(
            r[0] for r in conn.execute(
                "SELECT offline_order_no||'|'||online_order_no FROM matches WHERE event_id=?",
                (ev["id"],)
            ).fetchall()
        )

        new_matches = []
        for c in candidates:
            key = c["off_no"] + "|" + c["on_no"]
            if key in existing:
                continue
            existing.add(key)

            is_cancel = (c["off_status"] == "취소" and c["cancel_type"] == "pure_cancel")
            is_delivered = (
                c["on_status"] in DELIVERY_DONE_STATUSES
                or bool(c["on_delivery"] and c["on_delivery"] <= today_str)
            )

            if is_cancel and is_delivered:
                result_type = "cancel_match_delivered"
                compensation = round((c["off_amount"] or 0) * 0.05, 0)
            elif is_cancel:
                result_type = "cancel_match_pending"
                compensation = 0
            else:
                result_type = "active_match"
                compensation = 0

            phone_match = bool(c["off_phone"] and c["on_phone"] and c["off_phone"] == c["on_phone"])
            confidence = "high" if phone_match else "medium"
            keys = json.dumps(
                ["customer_name", "address_dong", "phone_last4"] if phone_match
                else ["customer_name", "address_dong"]
            )
            new_matches.append((c["off_no"], c["on_no"], ev["id"], keys, confidence, result_type, compensation))

        if new_matches:
            conn.executemany("""
                INSERT INTO matches
                (offline_order_no, online_order_no, event_id, match_keys,
                 match_confidence, result_type, compensation)
                VALUES (?,?,?,?,?,?,?)
                ON CONFLICT (offline_order_no, online_order_no) DO NOTHING
            """, new_matches)
            total_matched += len(new_matches)

    return {"matched": total_matched}


def auto_product_match(conn, match_id: int) -> dict:
    """
    erp_order_lines 기반 자동 품목 매칭 (3단계)
    - approved   : 시리즈+품목+세부품목1,2,3 완전 일치
    - manual_review: 시리즈+품목+세부품목1 일치, 세부품목2/3 다름
    - rejected   : 불일치 또는 제외 규칙 적용
    """
    match = conn.execute("SELECT * FROM matches WHERE id=?", (match_id,)).fetchone()
    if not match:
        return {"approved": 0, "manual_review": 0, "rejected": 0}

    offline_lines = conn.execute(
        "SELECT * FROM erp_order_lines WHERE order_no=?",
        (match["offline_order_no"],)
    ).fetchall()
    if not offline_lines:
        return {"approved": 0, "manual_review": 0, "rejected": 0}

    online_lines = conn.execute(
        "SELECT * FROM erp_order_lines WHERE order_no=?",
        (match["online_order_no"],)
    ).fetchall()
    have_online_detail = len(online_lines) > 0

    if not have_online_detail and match["event_id"]:
        event_prods = conn.execute(
            "SELECT series, category FROM event_products WHERE event_id=?",
            (match["event_id"],)
        ).fetchall()
        compare_list = [
            {"series": ep["series"], "category": ep["category"],
             "sub_product": None, "sub_product2": None, "sub_product3": None}
            for ep in event_prods
        ]
    else:
        compare_list = [dict(ol) for ol in online_lines]

    if not compare_list:
        return {"approved": 0, "manual_review": 0, "rejected": 0}

    conn.execute("DELETE FROM match_items WHERE match_id=?", (match_id,))

    approved_amount = 0
    cnt = {"approved": 0, "manual_review": 0, "rejected": 0}

    for ol in offline_lines:
        best = "rejected"
        for comp in compare_list:
            ol_series = ol["series"] or ""
            cp_series = comp.get("series") or ""
            if cp_series in EXCLUSION_RULES.get(ol_series, []):
                continue
            if ol_series in EXCLUSION_RULES.get(cp_series, []):
                continue
            if ol_series != cp_series:
                continue
            if (ol["category"] or "") != (comp.get("category") or ""):
                continue
            # 시리즈+품목 일치
            if not have_online_detail:
                best = "manual_review"
                break
            s1 = (ol["sub_product"] or "") == (comp.get("sub_product") or "")
            s2 = (ol["sub_product2"] or "") == (comp.get("sub_product2") or "")
            s3 = (ol["sub_product3"] or "") == (comp.get("sub_product3") or "")
            if s1 and s2 and s3:
                best = "approved"
                break
            elif s1 and best != "approved":
                best = "manual_review"

        cnt[best] += 1
        if best == "approved":
            approved_amount += ol["amount"] or 0

        conn.execute("""
            INSERT INTO match_items
            (match_id, product_name, series, category, sub_product, sub_product2, sub_product3, amount, match_status)
            VALUES (?,?,?,?,?,?,?,?,?)
        """, (match_id,
              ol["sub_product"] or "",
              ol["series"] or "",
              ol["category"] or "",
              ol["sub_product"] or "",
              ol["sub_product2"] or "",
              ol["sub_product3"] or "",
              ol["amount"] or 0,
              best))

    compensation = round(approved_amount * 0.05, 0)
    conn.execute("UPDATE matches SET compensation=? WHERE id=?", (compensation, match_id))
    return cnt


def refresh_delivery_status(conn) -> int:
    """
    기존 cancel_match_pending 건 중 온라인 수주가 납기완료된 건을
    cancel_match_delivered 로 업데이트하고 보전금 재계산
    매월 ERP 데이터 재업로드 후 호출
    """
    updated = 0
    pending_matches = conn.execute("""
        SELECT m.id, m.offline_order_no, m.online_order_no, of.amount
        FROM matches m
        JOIN erp_orders of ON m.offline_order_no = of.order_no
        WHERE m.result_type = 'cancel_match_pending'
          AND m.status = 'pending'
    """).fetchall()

    today = date.today().isoformat()
    for pm in pending_matches:
        online = conn.execute(
            "SELECT order_status, delivery_date FROM erp_orders WHERE order_no=?",
            (pm["online_order_no"],)
        ).fetchone()
        if not online:
            continue
        # 납기완료 상태 OR 확정납기일이 오늘 이전이면 납기 완료로 처리
        status_done = online["order_status"] in DELIVERY_DONE_STATUSES
        date_passed = bool(online["delivery_date"] and online["delivery_date"] <= today)
        if status_done or date_passed:
            compensation = round((pm["amount"] or 0) * 0.05, 0)
            conn.execute("""
                UPDATE matches
                SET result_type='cancel_match_delivered', compensation=?
                WHERE id=?
            """, (compensation, pm["id"]))
            updated += 1

    return updated


# ================================================================
# 기존 매칭 로직
# ================================================================
def is_same_series(s1: str, s2: str) -> bool:
    s1, s2 = s1.strip(), s2.strip()
    if s1 == s2:
        return True
    # 예외 규칙 적용
    if s2 in EXCLUSION_RULES.get(s1, []):
        return False
    return False


def match_products(order_products: list[dict], event_products: list[dict]) -> list[dict]:
    """수주 상품별 인정 여부 판정. event_products가 없으면 전체 인정."""
    results = []
    for op in order_products:
        if not event_products:
            matched = True
        else:
            matched = any(
                is_same_series(op["series"], ep["series"]) and op["category"] == ep["category"]
                for ep in event_products
            )
        results.append({**op, "match_status": "approved" if matched else "rejected"})
    return results


def calc_result(matched: list[dict], customer_amount: float) -> tuple[str, float, str]:
    """(result_type, compensation, note) 반환"""
    approved = [p for p in matched if p["match_status"] == "approved"]
    rejected = [p for p in matched if p["match_status"] == "rejected"]

    if not approved:
        return "rejected", 0.0, ""
    if not rejected:
        return "full", round(customer_amount * 0.05, 0), ""
    # 부분인정: 결제액 분리 기준이 정책에 없으므로 전체 5% 지급 후 별도 조정 안내
    note = f"부분인정 ({len(approved)}/{len(matched)}개 상품). 결제액 배분 기준 별도 확인 필요."
    return "partial", round(customer_amount * 0.05, 0), note


# ================================================================
# Pydantic Models
# ================================================================
class ProductIn(BaseModel):
    product_name: str
    series: str
    category: str


class EventIn(BaseModel):
    event_name: str
    announcement_date: str
    start_date: str
    end_date: str
    online_channel: str = "네이버"
    offline_lookback_days: int = 30
    products: List[ProductIn] = []


class OrderIn(BaseModel):
    store_name: str
    order_date: str
    customer_amount: float
    event_id: int
    products: List[ProductIn]


class MatchStatusUpdate(BaseModel):
    status: str


class ManualMatchIn(BaseModel):
    offline_order_no: str
    online_order_no: str
    event_id: int
    reason: str = ""


class MatchProductIn(BaseModel):
    product_name: str = ""
    series: str
    category: str
    amount: float = 0


def calc_result_by_product(matched: list[dict]) -> tuple[str, float, str]:
    """품목별 금액 기반 (result_type, compensation, note) 계산"""
    approved = [p for p in matched if p["match_status"] == "approved"]
    rejected = [p for p in matched if p["match_status"] == "rejected"]

    if not approved:
        return "rejected", 0.0, ""

    approved_amount = sum(p.get("amount", 0) for p in approved)
    compensation = round(approved_amount * 0.05, 0)

    if not rejected:
        return "full", compensation, ""

    note = (f"부분인정 ({len(approved)}/{len(matched)}개 상품) "
            f"| 인정 금액 ₩{approved_amount:,.0f}")
    return "partial", compensation, note


# ================================================================
# Routes
# ================================================================
@app.get("/", response_class=HTMLResponse)
async def home(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})


@app.get("/api/config")
async def get_config():
    return {"series": SERIES_LIST, "categories": CATEGORIES}


# --- 행사 ---
@app.post("/api/events")
async def create_event(event: EventIn):
    with get_db() as conn:
        # 커넥트플러스 기준: 동일 기간 행사 중복 방지
        overlapping = conn.execute("""
            SELECT id, event_name, start_date, end_date FROM events
            WHERE NOT (end_date < ? OR start_date > ?)
        """, (event.start_date, event.end_date)).fetchone()
        if overlapping:
            raise HTTPException(
                400,
                f"기간 중복: '{overlapping['event_name']}'({overlapping['start_date']}~{overlapping['end_date']})와 행사 기간이 겹칩니다. "
                f"커넥트플러스는 동시에 하나의 행사만 등록 가능합니다."
            )

        cur = conn.execute(
            "INSERT INTO events (event_name, announcement_date, start_date, end_date, online_channel, offline_lookback_days) VALUES (?,?,?,?,?,?) RETURNING id",
            (event.event_name, event.announcement_date, event.start_date, event.end_date, event.online_channel, event.offline_lookback_days),
        )
        event_id = cur.fetchone()["id"]
        for p in event.products:
            conn.execute(
                "INSERT INTO event_products (event_id, product_name, series, category) VALUES (?,?,?,?)",
                (event_id, p.product_name, p.series, p.category),
            )
    return {"id": event_id, "message": "행사 등록 완료"}


@app.get("/api/events")
async def list_events():
    with get_db() as conn:
        rows = conn.execute("""
            SELECT e.*,
                COUNT(m.id) FILTER (WHERE m.result_type='cancel_match_delivered')               AS confirmed_cnt,
                COUNT(m.id) FILTER (WHERE m.result_type='cancel_match_pending')                 AS pending_cnt,
                COUNT(m.id) FILTER (WHERE m.result_type='active_match')                         AS active_cnt,
                COUNT(m.id) FILTER (WHERE m.status='pending' AND m.result_type!='active_match') AS pending_review_cnt,
                COALESCE(SUM(m.compensation) FILTER (WHERE m.result_type='cancel_match_delivered'), 0) AS expected_compensation,
                COUNT(m.id) FILTER (WHERE m.status='approved')                                  AS approved_cnt,
                COALESCE(SUM(m.compensation) FILTER (WHERE m.status='approved'), 0)             AS approved_compensation
            FROM events e
            LEFT JOIN matches m ON e.id = m.event_id
            WHERE e.is_deleted IS NOT TRUE
            GROUP BY e.id
            ORDER BY e.announcement_date DESC
        """).fetchall()
        result = []
        for e in rows:
            ann = date.fromisoformat(e["announcement_date"])
            day_after = (ann + timedelta(days=1)).isoformat()
            lookback = e["offline_lookback_days"] if e["offline_lookback_days"] else 30
            offline_order_from = (ann - timedelta(days=lookback)).isoformat()
            ev_dict = dict(e)
            ev_dict["offline_extract_from"] = day_after
            ev_dict["offline_order_from"] = offline_order_from
            ev_dict["products"] = []
            result.append(ev_dict)
    return result


@app.get("/api/events/{event_id}/extract-guide")
async def get_extract_guide(event_id: int):
    """ERP 데이터 추출 가이드: 이 행사의 오프라인 취소 추출 기간 반환"""
    with get_db() as conn:
        ev = conn.execute("SELECT * FROM events WHERE id=?", (event_id,)).fetchone()
        if not ev:
            raise HTTPException(404, "행사를 찾을 수 없습니다")
        ann = date.fromisoformat(ev["announcement_date"])
        d_minus_1 = ann - timedelta(days=1)
        # 오프라인 취소 추출: 6개월 전 ~ D-1
        extract_from = (d_minus_1 - timedelta(days=180)).isoformat()
        return {
            "event_id": event_id,
            "event_name": ev["event_name"],
            "announcement_date": ev["announcement_date"],
            "online_period": f"{ev['start_date']} ~ {ev['end_date']}",
            "offline_extract_guide": {
                "description": "ERP에서 이 기간의 '오프라인 수주' 데이터를 추출하세요",
                "from": extract_from,
                "until": d_minus_1.isoformat(),
                "status_filter": "취소",
                "note": f"행사 공지일({ev['announcement_date']}) D-1({d_minus_1.isoformat()})까지 취소된 오프라인 수주만 보전 대상"
            }
        }


@app.patch("/api/events/{event_id}/close")
async def toggle_close_event(event_id: int):
    """행사 종결 토글 — 종결 ↔ 재개"""
    with get_db() as conn:
        ev = conn.execute("SELECT is_closed FROM events WHERE id=?", (event_id,)).fetchone()
        if not ev:
            raise HTTPException(404, "행사를 찾을 수 없습니다")
        new_state = not bool(ev["is_closed"])
        closed_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S") if new_state else None
        conn.execute(
            "UPDATE events SET is_closed=?, closed_at=? WHERE id=?",
            (new_state, closed_at, event_id),
        )
    return {"is_closed": new_state, "closed_at": closed_at}


@app.delete("/api/events/{event_id}")
async def delete_event(event_id: int):
    """소프트 삭제 — 실제 삭제 대신 is_deleted=TRUE 처리"""
    with get_db() as conn:
        ev = conn.execute("SELECT id FROM events WHERE id=? AND is_deleted IS NOT TRUE", (event_id,)).fetchone()
        if not ev:
            raise HTTPException(404, "행사를 찾을 수 없습니다")
        deleted_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        conn.execute("UPDATE events SET is_deleted=TRUE, deleted_at=? WHERE id=?", (deleted_at, event_id))
    return {"message": "삭제 완료"}


@app.get("/api/events/trash")
async def list_trash():
    """삭제된 행사 목록"""
    with get_db() as conn:
        rows = conn.execute(
            "SELECT id, event_name, start_date, end_date, online_channel, deleted_at FROM events WHERE is_deleted=TRUE ORDER BY deleted_at DESC"
        ).fetchall()
    return [dict(r) for r in rows]


@app.patch("/api/events/{event_id}/restore")
async def restore_event(event_id: int):
    """삭제된 행사 복원"""
    with get_db() as conn:
        ev = conn.execute("SELECT id FROM events WHERE id=? AND is_deleted=TRUE", (event_id,)).fetchone()
        if not ev:
            raise HTTPException(404, "복원할 행사를 찾을 수 없습니다")
        conn.execute("UPDATE events SET is_deleted=FALSE, deleted_at=NULL WHERE id=?", (event_id,))
    return {"message": "복원 완료"}


# 커넥트플러스 Excel 업로드
@app.post("/api/events/upload")
async def upload_events(file: UploadFile = File(...)):
    import openpyxl
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    # 헤더: 행사명 | 공지일 | 시작일 | 종료일 | 상품명 | 시리즈 | 품목
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    required = {"행사명", "공지일", "시작일", "종료일", "상품명", "시리즈", "품목"}
    if not required.issubset(set(headers)):
        raise HTTPException(400, f"Excel 헤더 오류. 필요: {required}, 실제: {set(headers)}")

    idx = {h: headers.index(h) for h in required}
    events_map: dict[str, dict] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[idx["행사명"]]:
            continue
        name = str(row[idx["행사명"]]).strip()
        if name not in events_map:
            events_map[name] = {
                "event_name": name,
                "announcement_date": str(row[idx["공지일"]])[:10],
                "start_date": str(row[idx["시작일"]])[:10],
                "end_date": str(row[idx["종료일"]])[:10],
                "products": [],
            }
        events_map[name]["products"].append(
            ProductIn(
                product_name=str(row[idx["상품명"]]).strip(),
                series=str(row[idx["시리즈"]]).strip(),
                category=str(row[idx["품목"]]).strip(),
            )
        )

    created = []
    for ev in events_map.values():
        res = await create_event(EventIn(**ev))
        created.append(res["id"])

    return {"message": f"{len(created)}개 행사 등록 완료", "event_ids": created}


# --- 수주 ---
@app.post("/api/orders")
async def create_order(order: OrderIn):
    with get_db() as conn:
        event = conn.execute("SELECT * FROM events WHERE id=?", (order.event_id,)).fetchone()
        if not event:
            raise HTTPException(404, "행사를 찾을 수 없습니다")

        # 수주 기간 체크: D-1 ~ 행사 마지막날
        od = date.fromisoformat(order.order_date)
        ad = date.fromisoformat(event["announcement_date"])
        d_minus_1 = ad - timedelta(days=1)
        event_end = date.fromisoformat(event["end_date"])
        if od < d_minus_1 or od > event_end:
            raise HTTPException(
                400,
                f"수주일({order.order_date})이 대상 기간(D-1: {d_minus_1} ~ 행사종료: {event_end}) 밖입니다. 매출기여 대상 외."
            )

        event_products = [
            dict(p) for p in conn.execute(
                "SELECT * FROM event_products WHERE event_id=?", (order.event_id,)
            ).fetchall()
        ]

        matched = match_products([p.model_dump() for p in order.products], event_products)
        result_type, compensation, note = calc_result(matched, order.customer_amount)

        cur = conn.execute(
            "INSERT INTO orders (store_name, order_date, customer_amount, event_id, result_type, compensation, note) VALUES (?,?,?,?,?,?,?) RETURNING id",
            (order.store_name, order.order_date, order.customer_amount,
             order.event_id, result_type, compensation, note),
        )
        order_id = cur.fetchone()["id"]

        for p in matched:
            conn.execute(
                "INSERT INTO order_products (order_id, product_name, series, category, match_status) VALUES (?,?,?,?,?)",
                (order_id, p["product_name"], p["series"], p["category"], p["match_status"]),
            )

    label = {"full": "✅ 전체 인정", "partial": "⚠️ 부분 인정", "rejected": "❌ 불인정"}
    return {
        "id": order_id,
        "result_type": result_type,
        "compensation": compensation,
        "products": matched,
        "note": note,
        "message": label[result_type],
    }


@app.get("/api/orders")
async def list_orders():
    with get_db() as conn:
        rows = conn.execute("""
            SELECT o.*, e.event_name
            FROM orders o
            LEFT JOIN events e ON o.event_id = e.id
            ORDER BY o.created_at DESC
        """).fetchall()
        result = []
        for o in rows:
            products = conn.execute(
                "SELECT * FROM order_products WHERE order_id=?", (o["id"],)
            ).fetchall()
            result.append({**dict(o), "products": [dict(p) for p in products]})
    return result


@app.delete("/api/orders/{order_id}")
async def delete_order(order_id: int):
    with get_db() as conn:
        conn.execute("DELETE FROM orders WHERE id=?", (order_id,))
    return {"message": "삭제 완료"}


@app.get("/api/stats")
async def get_stats():
    with get_db() as conn:
        s = conn.execute("""
            SELECT
                COUNT(*) as total,
                SUM(CASE WHEN result_type='full' THEN 1 ELSE 0 END) as full_cnt,
                SUM(CASE WHEN result_type='partial' THEN 1 ELSE 0 END) as partial_cnt,
                SUM(CASE WHEN result_type='rejected' THEN 1 ELSE 0 END) as rejected_cnt,
                SUM(COALESCE(compensation, 0)) as total_compensation
            FROM orders
        """).fetchone()
    return dict(s)


@app.get("/api/stats/by-store")
async def get_stats_by_store():
    """매장별 집계 통계"""
    with get_db() as conn:
        rows = conn.execute("""
            SELECT
                store_name,
                COUNT(*) as total,
                SUM(CASE WHEN result_type IN ('full','partial') THEN 1 ELSE 0 END) as approved_cnt,
                SUM(CASE WHEN result_type='rejected' THEN 1 ELSE 0 END) as rejected_cnt,
                SUM(COALESCE(compensation, 0)) as total_compensation
            FROM orders
            GROUP BY store_name
            ORDER BY total_compensation DESC
        """).fetchall()
    return [dict(r) for r in rows]


# ================================================================
# 엑셀 내보내기
# ================================================================
@app.get("/api/orders/export")
async def export_orders():
    """보전금 현황 엑셀 다운로드 (전체 수주 목록 + 매장별 집계)"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    BRAND_RED = "C80A1E"
    HEADER_BG = "C80A1E"
    HEADER_FONT = "FFFFFF"
    LIGHT_GRAY = "F5F5F5"
    GREEN_BG = "E8F5E9"
    YELLOW_BG = "FFF9C4"
    RED_BG = "FFEBEE"

    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def header_style(cell, bold=True):
        cell.font = Font(name="맑은 고딕", bold=bold, color=HEADER_FONT, size=10)
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    def data_style(cell, bold=False, color=None, align="left"):
        cell.font = Font(name="맑은 고딕", bold=bold, size=9)
        if color:
            cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border = border

    with get_db() as conn:
        orders = conn.execute("""
            SELECT o.*, e.event_name
            FROM orders o
            LEFT JOIN events e ON o.event_id = e.id
            ORDER BY o.store_name, o.order_date
        """).fetchall()
        store_stats = conn.execute("""
            SELECT
                store_name,
                COUNT(*) as total,
                SUM(CASE WHEN result_type IN ('full','partial') THEN 1 ELSE 0 END) as approved_cnt,
                SUM(CASE WHEN result_type='rejected' THEN 1 ELSE 0 END) as rejected_cnt,
                SUM(COALESCE(compensation, 0)) as total_compensation
            FROM orders
            GROUP BY store_name
            ORDER BY total_compensation DESC
        """).fetchall()

    wb = openpyxl.Workbook()

    # ── 시트 1: 전체 수주 목록 ──
    ws1 = wb.active
    ws1.title = "보전금_수주목록"
    ws1.sheet_view.showGridLines = False
    ws1.freeze_panes = "A2"

    headers1 = ["No", "매장명", "수주일", "행사명", "고객결제액(원)", "보전금(원)", "결과", "비고"]
    col_widths1 = [5, 14, 12, 24, 16, 14, 10, 30]
    ws1.row_dimensions[1].height = 22

    for col, (h, w) in enumerate(zip(headers1, col_widths1), 1):
        cell = ws1.cell(row=1, column=col, value=h)
        header_style(cell)
        ws1.column_dimensions[get_column_letter(col)].width = w

    result_map = {"full": "전체인정", "partial": "부분인정", "rejected": "불인정"}
    result_color = {"full": GREEN_BG, "partial": YELLOW_BG, "rejected": RED_BG}

    for i, o in enumerate(orders, 1):
        row = i + 1
        ws1.row_dimensions[row].height = 18
        rt = o["result_type"] or ""
        row_color = result_color.get(rt, None)

        values = [
            i,
            o["store_name"],
            o["order_date"],
            o["event_name"] or "-",
            o["customer_amount"],
            o["compensation"] or 0,
            result_map.get(rt, "-"),
            o["note"] or "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws1.cell(row=row, column=col, value=val)
            is_amount = col in (5, 6)
            data_style(cell, align="right" if is_amount else ("center" if col in (1, 3, 7) else "left"),
                       color=row_color if not is_amount else None)
            if is_amount:
                cell.number_format = '#,##0'
                if row_color:
                    cell.fill = PatternFill("solid", fgColor=row_color)

    # 합계 행
    total_row = len(orders) + 2
    ws1.row_dimensions[total_row].height = 20
    ws1.cell(total_row, 1, "합계").font = Font(name="맑은 고딕", bold=True, size=10)
    ws1.cell(total_row, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws1.cell(total_row, 1).fill = PatternFill("solid", fgColor="F0F0F0")
    ws1.cell(total_row, 1).border = border

    total_amount = sum(o["customer_amount"] for o in orders)
    total_comp = sum((o["compensation"] or 0) for o in orders)
    for col in range(2, 9):
        cell = ws1.cell(total_row, col)
        if col == 5:
            cell.value = total_amount
            cell.number_format = '#,##0'
            cell.font = Font(name="맑은 고딕", bold=True, size=10)
        elif col == 6:
            cell.value = total_comp
            cell.number_format = '#,##0'
            cell.font = Font(name="맑은 고딕", bold=True, color=BRAND_RED, size=10)
        cell.fill = PatternFill("solid", fgColor="F0F0F0")
        cell.alignment = Alignment(horizontal="right" if col in (5, 6) else "center", vertical="center")
        cell.border = border

    # ── 시트 2: 매장별 집계 ──
    ws2 = wb.create_sheet("매장별집계")
    ws2.sheet_view.showGridLines = False

    # 타이틀
    ws2.merge_cells("A1:E1")
    title_cell = ws2["A1"]
    title_cell.value = "일룸 매출기여 보전 프로그램 — 매장별 집계"
    title_cell.font = Font(name="맑은 고딕", bold=True, size=13, color=BRAND_RED)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    # 부제목
    ws2.merge_cells("A2:E2")
    sub_cell = ws2["A2"]
    sub_cell.value = f"출력일: {date.today().strftime('%Y-%m-%d')}  |  적용기간: 2026.05.01 ~ 2027.05.31"
    sub_cell.font = Font(name="맑은 고딕", size=9, color="888888")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[2].height = 16

    headers2 = ["매장명", "전체건수", "인정건수", "불인정건수", "보전금합계(원)"]
    col_widths2 = [16, 12, 12, 12, 18]
    ws2.row_dimensions[3].height = 22

    for col, (h, w) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=3, column=col, value=h)
        header_style(cell)
        ws2.column_dimensions[get_column_letter(col)].width = w

    for i, s in enumerate(store_stats, 1):
        row = i + 3
        ws2.row_dimensions[row].height = 18
        bg = LIGHT_GRAY if i % 2 == 0 else "FFFFFF"
        vals = [s["store_name"], s["total"], s["approved_cnt"], s["rejected_cnt"], s["total_compensation"]]
        for col, val in enumerate(vals, 1):
            cell = ws2.cell(row=row, column=col, value=val)
            is_amount = col == 5
            data_style(cell, align="right" if is_amount else ("center" if col > 1 else "left"), color=bg)
            if is_amount:
                cell.number_format = '#,##0'
                if s["total_compensation"] > 0:
                    cell.font = Font(name="맑은 고딕", bold=True, color=BRAND_RED, size=9)
                cell.fill = PatternFill("solid", fgColor=bg)

    # 합계 행
    total_row2 = len(store_stats) + 4
    ws2.row_dimensions[total_row2].height = 20
    totals2 = [
        "합계",
        sum(s["total"] for s in store_stats),
        sum(s["approved_cnt"] for s in store_stats),
        sum(s["rejected_cnt"] for s in store_stats),
        sum(s["total_compensation"] for s in store_stats),
    ]
    for col, val in enumerate(totals2, 1):
        cell = ws2.cell(total_row2, col, val)
        cell.font = Font(name="맑은 고딕", bold=True, size=10,
                         color=BRAND_RED if col == 5 else "333333")
        cell.fill = PatternFill("solid", fgColor="F0F0F0")
        cell.alignment = Alignment(
            horizontal="right" if col == 5 else ("center" if col > 1 else "left"),
            vertical="center"
        )
        cell.border = border
        if col == 5:
            cell.number_format = '#,##0'

    # 저장
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from urllib.parse import quote
    filename_raw = f"iloom_보전금현황_{date.today().strftime('%Y%m%d')}.xlsx"
    filename_encoded = quote(filename_raw, safe="")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}"}
    )


# ================================================================
# 태블로 수주 API
# ================================================================
@app.post("/api/tableau/upload")
async def upload_tableau(file: UploadFile = File(...)):
    """태블로 엑셀 업로드 → 수주번호 기준 집계 → 자동 매칭"""
    import openpyxl
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    headers_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(c).strip() if c is not None else '' for c in headers_row]

    required = ['수주번호', '수주건명', '수주상태', '실적대리점', '주문일자일', '동', '핸드폰번호']
    missing = [h for h in required if h not in headers]
    if missing:
        raise HTTPException(400, f"필수 컬럼 없음: {missing}\n실제 컬럼: {headers}")

    def ci(name):
        try: return headers.index(name)
        except ValueError: return None

    i_no = ci('수주번호')
    i_name = ci('수주건명')
    i_status = ci('수주상태')
    i_store = ci('실적대리점')
    i_date = ci('주문일자일')
    i_delivery = ci('확정납기일')
    i_dong = ci('동')
    i_phone = ci('핸드폰번호')
    i_setcode = ci('세트코드')
    i_series = ci('시리즈')
    i_category = ci('품목')
    i_sub1 = ci('세부품목')
    i_sub2 = ci('세부품목2')
    i_sub3 = ci('세부품목3')
    i_amount = len(headers) - 1  # 빈 헤더 마지막 컬럼 = 금액

    orders_map: dict[str, dict] = {}
    products_map: dict[str, list] = {}

    last_order_no = None          # 병합 셀 연속 행 처리용
    direct_skipped_nos: set = set()  # 직영매장 수주번호 (연속 행도 스킵)

    for row in ws.iter_rows(min_row=2, values_only=True):  # read_only 스트리밍
        order_no_raw = str(row[i_no]).strip() if row[i_no] else ''
        if order_no_raw and order_no_raw != 'None':
            # 새 수주 시작 행
            last_order_no = order_no_raw
            order_no = order_no_raw
        else:
            # 병합 셀 연속 행 → 직전 수주번호 승계
            order_no = last_order_no

        if not order_no:
            continue

        # 직영매장 수주 건너뜀 (연속 행 포함)
        if order_no in direct_skipped_nos:
            continue

        store = str(row[i_store]).strip() if row[i_store] else ''

        if order_no not in orders_map:
            # 첫 행: 직영 여부 확인 및 수주 등록
            if store in DIRECT_STORES:
                direct_skipped_nos.add(order_no)
                continue
            store_type = 'online' if store in ONLINE_STORES else 'offline'
            base, seq = parse_order_no(order_no)
            orders_map[order_no] = {
                'order_no': order_no,
                'order_base': base,
                'order_seq': seq,
                'store_name': store,
                'store_type': store_type,
                'order_status': str(row[i_status]).strip() if row[i_status] else '',
                'customer_name': extract_customer_name_tableau(row[i_name]),
                'order_date': parse_korean_date(row[i_date]),
                'delivery_date': parse_korean_date(row[i_delivery]) if i_delivery is not None else None,
                'address_dong': str(row[i_dong]).strip() if row[i_dong] else None,
                'phone_last4': extract_phone_last4_tableau(row[i_phone]) if i_phone is not None else None,
                'amount': 0,
            }
            products_map[order_no] = []

        # 금액 및 품목 누적 (첫 행 + 연속 행 모두)
        amt = float(row[i_amount]) if (i_amount < len(row) and row[i_amount]) else 0.0
        orders_map[order_no]['amount'] += amt

        series = str(row[i_series]).strip() if (i_series is not None and row[i_series]) else ''
        category = str(row[i_category]).strip() if (i_category is not None and row[i_category]) else ''
        if series or category:
            products_map[order_no].append({
                'set_code': str(row[i_setcode]).strip() if (i_setcode is not None and row[i_setcode]) else '',
                'series': series,
                'category': category,
                'sub1': str(row[i_sub1]).strip() if (i_sub1 is not None and row[i_sub1]) else '',
                'sub2': str(row[i_sub2]).strip() if (i_sub2 is not None and row[i_sub2]) else '',
                'sub3': str(row[i_sub3]).strip() if (i_sub3 is not None and row[i_sub3]) else '',
                'amount': amt,
            })

    wb.close()
    inserted = skipped = 0
    with get_db() as conn:
        # 검토 완료(승인/반려) 상태 백업 — 재업로드 후 복원
        reviewed_rows = conn.execute("""
            SELECT id, offline_order_no, online_order_no, event_id, status, compensation
            FROM matches WHERE status IN ('approved', 'rejected')
        """).fetchall()
        reviewed_backup = {}
        for rm in reviewed_rows:
            items = conn.execute(
                "SELECT product_name, series, category, sub_product, sub_product2, sub_product3, amount, match_status "
                "FROM match_items WHERE match_id=?", (rm["id"],)
            ).fetchall()
            key = (rm["offline_order_no"], rm["online_order_no"])
            reviewed_backup[key] = {
                "status": rm["status"],
                "compensation": rm["compensation"],
                "items": [dict(i) for i in items],
            }

        # 이전 업로드 데이터 전체 초기화 (직영 잔여 데이터 제거)
        conn.execute("DELETE FROM matches")
        conn.execute("DELETE FROM erp_order_lines")
        conn.execute("DELETE FROM erp_orders")
        order_rows = []
        line_rows = []
        order_nos = []
        for order in orders_map.values():
            order_rows.append((
                order['order_no'], order['order_base'], order['order_seq'],
                order['store_name'], order['store_type'], order['order_status'],
                order['customer_name'], order['order_date'], order['delivery_date'],
                order['amount'], order['address_dong'], order['phone_last4'],
            ))
            order_nos.append(order['order_no'])
            for p in products_map[order['order_no']]:
                line_rows.append((order['order_no'], p['set_code'], p['series'], p['category'],
                                  p['sub1'], p['sub2'], p['sub3'], p['amount']))
        conn.executemany("""
            INSERT INTO erp_orders
            (order_no, order_base, order_seq, store_name, store_type, order_status,
             customer_name, order_date, delivery_date, amount, address_dong, phone_last4)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT (order_no) DO UPDATE SET
                order_base=EXCLUDED.order_base, order_seq=EXCLUDED.order_seq,
                store_name=EXCLUDED.store_name, store_type=EXCLUDED.store_type,
                order_status=EXCLUDED.order_status, customer_name=EXCLUDED.customer_name,
                order_date=EXCLUDED.order_date, delivery_date=EXCLUDED.delivery_date,
                amount=EXCLUDED.amount, address_dong=EXCLUDED.address_dong,
                phone_last4=EXCLUDED.phone_last4
        """, order_rows)
        conn.execute("DELETE FROM erp_order_lines WHERE order_no = ANY(?)", (list(order_nos),))
        conn.executemany("""
            INSERT INTO erp_order_lines
            (order_no, set_code, series, category, sub_product, sub_product2, sub_product3, amount)
            VALUES (?,?,?,?,?,?,?,?)
        """, line_rows)
        inserted = len(order_rows)

        classify_cancel_types(conn)
        match_result = run_matching_engine(conn)
        delivery_updated = refresh_delivery_status(conn)

        # 검토 완료 상태 복원
        restored = 0
        for (off_no, on_no), bak in reviewed_backup.items():
            m = conn.execute(
                "SELECT id FROM matches WHERE offline_order_no=? AND online_order_no=?",
                (off_no, on_no)
            ).fetchone()
            if not m:
                continue
            conn.execute(
                "UPDATE matches SET status=?, compensation=? WHERE id=?",
                (bak["status"], bak["compensation"], m["id"])
            )
            if bak["items"]:
                conn.execute("DELETE FROM match_items WHERE match_id=?", (m["id"],))
                conn.executemany("""
                    INSERT INTO match_items
                    (match_id, product_name, series, category, sub_product, sub_product2, sub_product3, amount, match_status)
                    VALUES (?,?,?,?,?,?,?,?,?)
                """, [
                    (m["id"], it["product_name"], it["series"], it["category"],
                     it["sub_product"], it["sub_product2"], it["sub_product3"],
                     it["amount"], it["match_status"])
                    for it in bak["items"]
                ])
            restored += 1

        total_online = conn.execute("SELECT COUNT(*) as c FROM erp_orders WHERE store_type='online'").fetchone()["c"]
        total_offline = conn.execute("SELECT COUNT(*) as c FROM erp_orders WHERE store_type='offline'").fetchone()["c"]
        pure_cancel_cnt = conn.execute(
            "SELECT COUNT(*) as c FROM erp_orders WHERE store_type='offline' AND cancel_type='pure_cancel'"
        ).fetchone()["c"]
        high_conf = conn.execute("SELECT COUNT(*) as c FROM matches WHERE match_confidence='high'").fetchone()["c"]
        med_conf = conn.execute("SELECT COUNT(*) as c FROM matches WHERE match_confidence='medium'").fetchone()["c"]

    return {
        "message": f"{inserted}건 업로드 완료 (건너뜀 {skipped}건)",
        "uploaded": inserted,
        "skipped": skipped,
        "store_breakdown": {"online": total_online, "offline": total_offline},
        "cancel_summary": {"보전대상_오프라인_취소": pure_cancel_cnt},
        "matching": {
            "신규_매칭건수": match_result["matched"],
            "납기완료_업데이트": delivery_updated,
            "high_신뢰도_핸드폰일치": high_conf,
            "medium_신뢰도_이름동일치": med_conf,
            "검토상태_복원": restored,
        }
    }


# ================================================================
# ERP 수주 API
# ================================================================
@app.post("/api/erp/upload")
async def upload_erp(file: UploadFile = File(...)):
    """ERP 수주 Excel 업로드 → 자동 분류 + 매칭"""
    import openpyxl
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]

    required = ["대리점", "수주번호", "수주상태", "주문일자", "수주건별금액", "납품처주소", "수주건명"]
    missing = [h for h in required if h not in headers]
    if missing:
        raise HTTPException(400, f"필수 컬럼 없음: {missing}")

    def ci(name: str) -> int | None:
        try:
            return headers.index(name)
        except ValueError:
            return None

    i_store = ci("대리점")
    i_no = ci("수주번호")
    i_status = ci("수주상태")
    i_name = ci("수주건명")
    i_date = ci("주문일자")
    i_delivery = ci("확정납기")
    i_amount = ci("수주건별금액")
    i_addr = ci("납품처주소")
    i_biz = ci("사업소")  # 온라인사업부 여부 판별용

    def to_date(v) -> str | None:
        if v is None:
            return None
        if hasattr(v, "strftime"):
            return v.strftime("%Y-%m-%d")
        return str(v)[:10]

    inserted = skipped = 0
    with get_db() as conn:
        for row in ws.iter_rows(min_row=2, values_only=True):
            order_no = row[i_no]
            store = row[i_store]
            if not order_no or not store:
                continue

            order_no = str(order_no).strip()
            store = str(store).strip()
            status = str(row[i_status]).strip() if row[i_status] else ""
            order_name = str(row[i_name]).strip() if row[i_name] else ""
            customer_name = extract_customer_name(order_name)
            base, seq = parse_order_no(order_no)
            biz = str(row[i_biz]).strip() if (i_biz is not None and row[i_biz]) else ""
            store_type = "online" if (store in ONLINE_STORES or biz == "온라인사업부") else "offline"
            order_date = to_date(row[i_date])
            delivery_date = to_date(row[i_delivery]) if i_delivery is not None else None
            amount = float(row[i_amount]) if row[i_amount] else 0.0
            address = str(row[i_addr]).strip() if row[i_addr] else None

            try:
                conn.execute("""
                    INSERT INTO erp_orders
                    (order_no, order_base, order_seq, store_name, store_type, order_status,
                     customer_name, order_name, order_date, delivery_date, amount, address_dong)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                    ON CONFLICT (order_no) DO UPDATE SET
                        order_base=EXCLUDED.order_base, order_seq=EXCLUDED.order_seq,
                        store_name=EXCLUDED.store_name, store_type=EXCLUDED.store_type,
                        order_status=EXCLUDED.order_status, customer_name=EXCLUDED.customer_name,
                        order_name=EXCLUDED.order_name, order_date=EXCLUDED.order_date,
                        delivery_date=EXCLUDED.delivery_date, amount=EXCLUDED.amount,
                        address_dong=EXCLUDED.address_dong
                """, (order_no, base, seq, store, store_type, status,
                      customer_name, order_name, order_date, delivery_date, amount, address))
                inserted += 1
            except Exception:
                skipped += 1

        classify_cancel_types(conn)
        match_result = run_matching_engine(conn)
        matched = match_result["matched"]
        delivery_updated = refresh_delivery_status(conn)

        # 취소 분류 상세 통계
        cancel_stats = conn.execute("""
            SELECT
                cancel_type,
                COUNT(*) as cnt,
                SUM(amount) as total_amount
            FROM erp_orders
            WHERE order_status='취소'
            GROUP BY cancel_type
        """).fetchall()

        # 행사별 매칭 현황
        event_match_stats = conn.execute("""
            SELECT e.event_name, COUNT(m.id) as match_cnt,
                   SUM(m.compensation) as total_comp
            FROM events e
            LEFT JOIN matches m ON e.id = m.event_id
            GROUP BY e.id, e.event_name
        """).fetchall()

        pure_cancel_cnt = conn.execute(
            "SELECT COUNT(*) as c FROM erp_orders WHERE store_type='offline' AND cancel_type='pure_cancel'"
        ).fetchone()["c"]

    cancel_breakdown = {r["cancel_type"] or "미분류": {"건수": r["cnt"], "금액": r["total_amount"]} for r in cancel_stats}
    event_breakdown = [{"행사명": r["event_name"], "매칭건수": r["match_cnt"], "보전금합계": r["total_comp"] or 0} for r in event_match_stats]

    return {
        "message": f"{inserted}건 처리 완료",
        "inserted": inserted,
        "skipped": skipped,
        "cancel_summary": {
            "pure_cancel": cancel_breakdown.get("pure_cancel", {"건수": 0, "금액": 0}),
            "date_change": cancel_breakdown.get("date_change", {"건수": 0, "금액": 0}),
            "미분류": cancel_breakdown.get("미분류", {"건수": 0, "금액": 0}),
            "보전대상_오프라인_취소": pure_cancel_cnt,
        },
        "matching": {
            "신규_매칭건수": matched,
            "납기완료_업데이트": delivery_updated,
            "행사별_현황": event_breakdown,
        },
    }


@app.get("/api/erp/orders")
async def list_erp_orders(store_type: str = "", status: str = "", cancel_type: str = "", limit: int = 10):
    with get_db() as conn:
        q = "SELECT * FROM erp_orders WHERE 1=1"
        params: list = []
        if store_type:
            q += " AND store_type=?"; params.append(store_type)
        if status:
            q += " AND order_status=?"; params.append(status)
        if cancel_type:
            q += " AND cancel_type=?"; params.append(cancel_type)
        q += " ORDER BY order_date DESC, order_no LIMIT ?"
        params.append(min(limit, 50))
        rows = conn.execute(q, params).fetchall()
    return [dict(r) for r in rows]


@app.get("/api/erp/stats")
async def erp_stats():
    with get_db() as conn:
        cancel_cnt = conn.execute(
            "SELECT COUNT(*) as c FROM erp_orders WHERE store_type='offline' AND cancel_type='pure_cancel'"
        ).fetchone()["c"]
        match_cnt = conn.execute("SELECT COUNT(*) as c FROM matches").fetchone()["c"]
        delivered_cnt = conn.execute(
            "SELECT COUNT(*) as c FROM matches WHERE result_type='cancel_match_delivered'"
        ).fetchone()["c"]
        pending_cnt = conn.execute(
            "SELECT COUNT(*) as c FROM matches WHERE result_type='cancel_match_pending'"
        ).fetchone()["c"]
        approved_cnt = conn.execute(
            "SELECT COUNT(*) as c FROM matches WHERE status='approved'"
        ).fetchone()["c"]
        total_comp = conn.execute(
            "SELECT COALESCE(SUM(compensation),0) as s FROM matches WHERE status='approved'"
        ).fetchone()["s"]
        # 납기완료 기준 보전 예정액 (미승인 포함)
        expected_comp = conn.execute(
            "SELECT COALESCE(SUM(compensation),0) as s FROM matches WHERE result_type='cancel_match_delivered'"
        ).fetchone()["s"]
    return {
        "total_cancels": cancel_cnt,
        "total_matches": match_cnt,
        "delivered_matches": delivered_cnt,
        "pending_delivery_matches": pending_cnt,
        "approved_matches": approved_cnt,
        "total_compensation": total_comp,
        "expected_compensation": expected_comp,
    }


@app.post("/api/erp/run-match")
async def trigger_match():
    with get_db() as conn:
        matched = run_matching_engine(conn)
        updated = refresh_delivery_status(conn)
    return {
        "message": f"매칭 실행 완료",
        "신규_매칭": matched,
        "납기완료_업데이트": updated,
    }


@app.post("/api/erp/refresh-delivery")
async def refresh_delivery():
    """납기완료 상태 업데이트만 단독 실행 (매월 ERP 재업로드 후 호출)"""
    with get_db() as conn:
        updated = refresh_delivery_status(conn)
    return {"message": f"납기완료 업데이트: {updated}건 → cancel_match_delivered 전환"}


# ================================================================
# 매칭 결과 API
# ================================================================
@app.get("/api/matches")
async def list_matches():
    with get_db() as conn:
        rows = conn.execute("""
            SELECT
                m.id, m.offline_order_no, m.online_order_no,
                m.match_keys, m.match_confidence, m.result_type,
                m.compensation, m.status, m.created_at,
                off.store_name  AS offline_store,
                off.customer_name,
                off.address_dong,
                off.order_date  AS offline_date,
                off.amount      AS offline_amount,
                on_.store_name  AS online_store,
                on_.order_date  AS online_date,
                on_.amount      AS online_amount,
                e.event_name
            FROM matches m
            JOIN erp_orders off ON m.offline_order_no = off.order_no
            JOIN erp_orders on_ ON m.online_order_no  = on_.order_no
            LEFT JOIN events e  ON m.event_id = e.id
            ORDER BY m.created_at DESC
        """).fetchall()

        result = []
        for r in rows:
            d = dict(r)
            items = conn.execute(
                "SELECT * FROM match_items WHERE match_id=? ORDER BY id", (r["id"],)
            ).fetchall()
            d["items"] = [dict(i) for i in items]
            result.append(d)
    return result


@app.get("/api/clusters")
async def list_clusters(event_id: int = 0):
    """
    사람(고객) 기준 클러스터 뷰 (선우님 Step 4)
    동일 고객이 온라인+오프라인 모두에서 수주한 건들을 묶어서 반환
    """
    with get_db() as conn:
        where = "WHERE m.event_id = ?" if event_id else ""
        params = [event_id] if event_id else []

        rows = conn.execute(f"""
            SELECT
                of.customer_name,
                of.address_dong,
                e.id                                                                    AS event_id,
                e.event_name,
                COUNT(DISTINCT of.order_no)                                                     AS offline_order_cnt,
                SUM(CASE WHEN m.result_type='cancel_match_delivered' THEN 1 ELSE 0 END)        AS cancel_cnt,
                SUM(CASE WHEN m.result_type='cancel_match_pending'   THEN 1 ELSE 0 END)        AS pending_cnt,
                SUM(CASE WHEN m.result_type='active_match'           THEN 1 ELSE 0 END)        AS active_cnt,
                SUM(CASE WHEN m.result_type='cancel_match_delivered' THEN m.compensation ELSE 0 END) AS total_compensation,
                STRING_AGG(DISTINCT of.store_name, ',')                                AS offline_stores,
                STRING_AGG(DISTINCT on_.store_name, ',')                               AS online_stores,
                MIN(m.match_confidence)                                                AS min_confidence
            FROM matches m
            JOIN erp_orders of  ON m.offline_order_no = of.order_no
            JOIN erp_orders on_ ON m.online_order_no  = on_.order_no
            JOIN events e       ON m.event_id = e.id
            {where}
            GROUP BY of.customer_name, of.address_dong, e.id
            ORDER BY total_compensation DESC, of.customer_name
        """, params).fetchall()
    return [dict(r) for r in rows]


@app.get("/api/clusters/detail")
async def cluster_detail(event_id: int, customer: str, address: str):
    """특정 고객+행사의 온라인+오프라인 수주 상세 목록"""
    with get_db() as conn:
        rows = conn.execute("""
            SELECT
                m.id, m.result_type, m.compensation, m.match_confidence, m.status,
                of.order_no       AS offline_order_no,
                of.store_name     AS offline_store,
                of.order_date     AS offline_date,
                of.delivery_date  AS offline_delivery_date,
                of.amount         AS offline_amount,
                of.order_status   AS offline_status,
                of.cancel_type,
                on_.order_no      AS online_order_no,
                on_.store_name    AS online_store,
                on_.order_date    AS online_date,
                on_.delivery_date AS online_delivery_date,
                on_.amount        AS online_amount,
                e.event_name
            FROM matches m
            JOIN erp_orders of  ON m.offline_order_no = of.order_no
            JOIN erp_orders on_ ON m.online_order_no  = on_.order_no
            LEFT JOIN events e  ON m.event_id = e.id
            WHERE m.event_id = ?
            AND of.customer_name = ?
            AND of.address_dong = ?
            ORDER BY m.result_type DESC, m.id
        """, (event_id, customer, address)).fetchall()
        result = []
        for r in rows:
            d = dict(r)
            items = conn.execute(
                "SELECT * FROM match_items WHERE match_id=? ORDER BY id", (r["id"],)
            ).fetchall()
            d["items"] = [dict(i) for i in items]
            result.append(d)
    return result


@app.post("/api/matches/{match_id}/products")
async def verify_match_products(match_id: int, products: List[MatchProductIn]):
    """오프라인 취소 품목 입력 → 행사 품목과 매칭 → 보전금 계산"""
    if not products:
        raise HTTPException(400, "품목을 1개 이상 입력하세요")

    with get_db() as conn:
        match = conn.execute("SELECT * FROM matches WHERE id=?", (match_id,)).fetchone()
        if not match:
            raise HTTPException(404, "매칭 건을 찾을 수 없습니다")

        # 행사 품목 조회
        event_products = []
        if match["event_id"]:
            event_products = [
                dict(r) for r in conn.execute(
                    "SELECT * FROM event_products WHERE event_id=?", (match["event_id"],)
                ).fetchall()
            ]

        # 기존 품목 삭제 후 재저장
        conn.execute("DELETE FROM match_items WHERE match_id=?", (match_id,))

        product_dicts = [p.model_dump() for p in products]

        if event_products:
            matched = match_products(product_dicts, event_products)
        else:
            # 행사 미등록 시 전체 인정 처리 (수동 확인 필요 표시)
            matched = [{**p, "match_status": "approved"} for p in product_dicts]

        for mp in matched:
            conn.execute("""
                INSERT INTO match_items (match_id, product_name, series, category, amount, match_status)
                VALUES (?,?,?,?,?,?)
            """, (match_id, mp.get("product_name", ""), mp["series"],
                  mp["category"], mp.get("amount", 0), mp["match_status"]))

        result_type, compensation, note = calc_result_by_product(matched)

        conn.execute("""
            UPDATE matches SET result_type=?, compensation=?
            WHERE id=?
        """, (result_type, compensation, match_id))

    label = {"full": "✅ 전체 인정", "partial": "⚠️ 부분 인정", "rejected": "❌ 불인정"}
    return {
        "result_type": result_type,
        "compensation": compensation,
        "note": note,
        "products": matched,
        "message": label.get(result_type, "-"),
    }


@app.post("/api/matches/{match_id}/auto-product-match")
async def trigger_auto_product_match(match_id: int):
    with get_db() as conn:
        cnt = auto_product_match(conn, match_id)
    return {"approved": cnt["approved"], "manual_review": cnt["manual_review"], "rejected": cnt["rejected"]}


@app.post("/api/events/{event_id}/auto-match-all")
async def auto_match_all(event_id: int):
    """해당 행사의 검토대기 매칭 전체 자동 품목매칭 일괄 실행"""
    with get_db() as conn:
        pending_ids = [
            r["id"] for r in conn.execute(
                "SELECT id FROM matches WHERE event_id=? AND status='pending'",
                (event_id,)
            ).fetchall()
        ]
        totals = {"approved": 0, "manual_review": 0, "rejected": 0}
        for mid in pending_ids:
            cnt = auto_product_match(conn, mid)
            for k in totals:
                totals[k] += cnt[k]
    return {
        "message": f"{len(pending_ids)}건 자동매칭 완료",
        "processed": len(pending_ids),
        **totals,
    }


@app.patch("/api/match-items/{item_id}/decide")
async def decide_match_item(item_id: int, body: dict = Body(...)):
    decision = body.get("status")
    if decision not in ("approved", "rejected"):
        raise HTTPException(400, "status must be 'approved' or 'rejected'")
    with get_db() as conn:
        conn.execute("UPDATE match_items SET match_status=? WHERE id=?", (decision, item_id))
        row = conn.execute("SELECT match_id FROM match_items WHERE id=?", (item_id,)).fetchone()
        if row:
            approved_amount = conn.execute(
                "SELECT COALESCE(SUM(amount),0) as s FROM match_items WHERE match_id=? AND match_status='approved'",
                (row["match_id"],)
            ).fetchone()["s"]
            conn.execute("UPDATE matches SET compensation=? WHERE id=?",
                         (round(approved_amount * 0.05, 0), row["match_id"]))
    return {"ok": True}


@app.patch("/api/matches/{match_id}")
async def update_match(match_id: int, body: MatchStatusUpdate):
    if body.status not in ("pending", "approved", "rejected"):
        raise HTTPException(400, "status must be pending / approved / rejected")
    with get_db() as conn:
        conn.execute("UPDATE matches SET status=? WHERE id=?", (body.status, match_id))
    return {"message": "상태 업데이트 완료"}


@app.post("/api/matches/manual")
async def create_manual_match(body: ManualMatchIn):
    """수기 매칭 등록 — 담당자가 호수까지 확인한 동명이인/가족 명의 예외 처리"""
    with get_db() as conn:
        offline = conn.execute(
            "SELECT * FROM erp_orders WHERE order_no=?", (body.offline_order_no,)
        ).fetchone()
        if not offline:
            raise HTTPException(404, f"오프라인 수주번호를 찾을 수 없습니다: {body.offline_order_no}")

        online = conn.execute(
            "SELECT * FROM erp_orders WHERE order_no=?", (body.online_order_no,)
        ).fetchone()
        if not online:
            raise HTTPException(404, f"온라인 수주번호를 찾을 수 없습니다: {body.online_order_no}")

        exists = conn.execute(
            "SELECT 1 FROM matches WHERE offline_order_no=? AND online_order_no=?",
            (body.offline_order_no, body.online_order_no)
        ).fetchone()
        if exists:
            raise HTTPException(400, "이미 매칭된 건입니다")

        is_cancel = (
            offline["order_status"] == "취소"
            and offline["cancel_type"] == "pure_cancel"
        )
        is_delivered = online["order_status"] in DELIVERY_DONE_STATUSES

        if is_cancel and is_delivered:
            result_type = "cancel_match_delivered"
            compensation = round((offline["amount"] or 0) * 0.05, 0)
        elif is_cancel:
            result_type = "cancel_match_pending"
            compensation = 0
        else:
            result_type = "active_match"
            compensation = 0

        keys = json.dumps(["manual", body.reason] if body.reason else ["manual"])
        conn.execute("""
            INSERT INTO matches
            (offline_order_no, online_order_no, event_id, match_keys,
             match_confidence, result_type, compensation)
            VALUES (?,?,?,?,?,?,?)
        """, (
            body.offline_order_no, body.online_order_no, body.event_id,
            keys, "manual", result_type, compensation,
        ))

    label = {"cancel_match_delivered": "✅ 보전 확정", "cancel_match_pending": "⏳ 보전 대기", "active_match": "모니터링"}
    return {"message": f"수기 매칭 등록 완료 — {label.get(result_type, result_type)}"}


@app.get("/api/events/{event_id}/stats")
async def event_stats(event_id: int):
    with get_db() as conn:
        s = conn.execute("""
            SELECT
                COUNT(*) as total_matches,
                SUM(CASE WHEN result_type='cancel_match_delivered' THEN 1 ELSE 0 END) as confirmed_cnt,
                SUM(CASE WHEN result_type='cancel_match_pending'   THEN 1 ELSE 0 END) as pending_cnt,
                SUM(CASE WHEN result_type='active_match'           THEN 1 ELSE 0 END) as active_cnt,
                SUM(CASE WHEN status='pending' AND result_type != 'active_match' THEN 1 ELSE 0 END) as pending_review_cnt,
                SUM(CASE WHEN result_type='cancel_match_delivered' THEN compensation ELSE 0 END) as expected_compensation,
                SUM(CASE WHEN status='approved' THEN 1 ELSE 0 END) as approved_cnt,
                SUM(CASE WHEN status='approved' THEN compensation ELSE 0 END) as approved_compensation,
                SUM(CASE WHEN status='rejected' THEN 1 ELSE 0 END) as rejected_cnt
            FROM matches WHERE event_id=?
        """, (event_id,)).fetchone()
        cancel_cnt = conn.execute(
            "SELECT COUNT(*) as c FROM erp_orders WHERE store_type='offline' AND cancel_type='pure_cancel'"
        ).fetchone()["c"]
    return {**dict(s), "offline_cancel_cnt": cancel_cnt}


@app.get("/api/events/{event_id}/store-stats")
async def event_store_stats(event_id: int):
    with get_db() as conn:
        rows = conn.execute("""
            SELECT
                off.store_name,
                COUNT(*) as total,
                SUM(CASE WHEN m.status='approved' THEN 1 ELSE 0 END) as approved_cnt,
                SUM(CASE WHEN m.status='rejected' THEN 1 ELSE 0 END) as rejected_cnt,
                SUM(CASE WHEN m.result_type='cancel_match_pending' THEN 1 ELSE 0 END) as pending_cnt,
                SUM(CASE WHEN m.result_type='cancel_match_delivered' THEN m.compensation ELSE 0 END) as expected_compensation,
                SUM(CASE WHEN m.status='approved' THEN m.compensation ELSE 0 END) as total_compensation
            FROM matches m
            JOIN erp_orders off ON m.offline_order_no = off.order_no
            WHERE m.event_id=? AND m.result_type != 'active_match'
            GROUP BY off.store_name
            ORDER BY expected_compensation DESC
        """, (event_id,)).fetchall()
    return [dict(r) for r in rows]


@app.get("/api/events/{event_id}/export")
async def export_event(event_id: int):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    BRAND_RED = "C80A1E"
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hstyle(cell):
        cell.font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=BRAND_RED)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    def dstyle(cell, bold=False, color=None, align="left"):
        cell.font = Font(name="맑은 고딕", bold=bold, size=9)
        if color:
            cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border = border

    with get_db() as conn:
        ev = conn.execute("SELECT * FROM events WHERE id=?", (event_id,)).fetchone()
        if not ev:
            raise HTTPException(404, "행사를 찾을 수 없습니다")

        rows = conn.execute("""
            SELECT
                off.store_name, off.customer_name, off.address_dong,
                off.order_date AS offline_date, off.amount AS offline_amount,
                on_.order_date AS online_date, on_.amount AS online_amount,
                on_.store_name AS online_store,
                m.match_confidence, m.result_type, m.compensation, m.status,
                e.event_name
            FROM matches m
            JOIN erp_orders off ON m.offline_order_no = off.order_no
            JOIN erp_orders on_ ON m.online_order_no  = on_.order_no
            LEFT JOIN events e  ON m.event_id = e.id
            WHERE m.event_id=? AND m.result_type != 'active_match'
            ORDER BY off.store_name, m.result_type DESC
        """, (event_id,)).fetchall()

        store_stats = conn.execute("""
            SELECT
                off.store_name,
                COUNT(*) as total,
                SUM(CASE WHEN m.status='approved' THEN 1 ELSE 0 END) as approved_cnt,
                SUM(CASE WHEN m.result_type='cancel_match_delivered' THEN m.compensation ELSE 0 END) as expected_compensation,
                SUM(CASE WHEN m.status='approved' THEN m.compensation ELSE 0 END) as total_compensation
            FROM matches m
            JOIN erp_orders off ON m.offline_order_no = off.order_no
            WHERE m.event_id=? AND m.result_type != 'active_match'
            GROUP BY off.store_name ORDER BY expected_compensation DESC
        """, (event_id,)).fetchall()

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "보전금_상세"
    ws1.sheet_view.showGridLines = False
    ws1.freeze_panes = "A2"

    h1 = ["No", "매장명", "고객명", "동주소", "오프라인 수주일", "오프라인 금액", "온라인 수주일", "온라인 금액", "신뢰도", "결과", "보전금(원)", "처리상태"]
    cw1 = [5, 14, 10, 12, 14, 14, 14, 14, 10, 10, 14, 10]
    ws1.row_dimensions[1].height = 22
    for col, (h, w) in enumerate(zip(h1, cw1), 1):
        hstyle(ws1.cell(1, col, h))
        ws1.column_dimensions[get_column_letter(col)].width = w

    rt_label = {"cancel_match_delivered": "보전확정", "cancel_match_pending": "납기대기"}
    st_label = {"approved": "승인", "rejected": "반려", "pending": "검토중"}
    conf_label = {"high": "HIGH", "medium": "MEDIUM", "manual": "수기"}
    LIGHT = "F5F5F5"

    for i, r in enumerate(rows, 1):
        row = i + 1
        ws1.row_dimensions[row].height = 18
        bg = LIGHT if i % 2 == 0 else "FFFFFF"
        vals = [
            i, r["store_name"], r["customer_name"], r["address_dong"] or "-",
            r["offline_date"], r["offline_amount"] or 0,
            r["online_date"], r["online_amount"] or 0,
            conf_label.get(r["match_confidence"], "-"),
            rt_label.get(r["result_type"], "-"),
            r["compensation"] or 0,
            st_label.get(r["status"], "-"),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws1.cell(row, col, val)
            is_amt = col in (6, 8, 11)
            dstyle(cell, color=bg, align="right" if is_amt else ("center" if col in (1, 9, 10, 12) else "left"))
            if is_amt:
                cell.number_format = '#,##0'
                cell.fill = PatternFill("solid", fgColor=bg)

    ws2 = wb.create_sheet("매장별집계")
    ws2.sheet_view.showGridLines = False
    h2 = ["매장명", "전체건수", "승인건수", "보전 예정액(원)", "확정 보전금(원)"]
    cw2 = [16, 12, 12, 18, 18]
    ws2.row_dimensions[1].height = 22
    for col, (h, w) in enumerate(zip(h2, cw2), 1):
        hstyle(ws2.cell(1, col, h))
        ws2.column_dimensions[get_column_letter(col)].width = w

    for i, s in enumerate(store_stats, 1):
        row = i + 1
        bg = LIGHT if i % 2 == 0 else "FFFFFF"
        vals = [s["store_name"], s["total"], s["approved_cnt"], s["expected_compensation"] or 0, s["total_compensation"] or 0]
        for col, val in enumerate(vals, 1):
            cell = ws2.cell(row, col, val)
            dstyle(cell, color=bg, align="right" if col >= 4 else ("center" if col > 1 else "left"))
            if col >= 4:
                cell.number_format = '#,##0'
                cell.fill = PatternFill("solid", fgColor=bg)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from urllib.parse import quote
    event_name = ev["event_name"]
    fname = quote(f"iloom_보전금_{event_name}_{date.today().strftime('%Y%m%d')}.xlsx", safe="")
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{fname}"})


if __name__ == "__main__":
    uvicorn.run("app:app", host="0.0.0.0", port=8000, reload=True)
