import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "송도점 퇴점 현황"

# ── 컬러 팔레트 ──────────────────────────────
C_TITLE    = "2C2C2C"
C_SUBTITLE = "555555"
C_SEC_HDR  = "4A4A4A"
C_COL_HDR  = "6B6B6B"
C_LABEL    = "E4E4E4"
C_BG_EVEN  = "F8F8F8"
C_BG_ODD   = "FFFFFF"
C_EMPH_BG  = "EBEBEB"
C_DARK     = "333333"
C_WHITE    = "FFFFFF"
C_GRAY_BDR = "C8C8C8"
C_EMPH_TXT = "1A1A1A"
C_LINK     = "1155CC"   # 링크 파란색

def bd():
    s = Side(style="thin", color=C_GRAY_BDR)
    return Border(left=s, right=s, top=s, bottom=s)

def fill(h): return PatternFill("solid", fgColor=h)
def rh(r, h): ws.row_dimensions[r].height = h

def fnt(bold=False, size=10, color=C_DARK, italic=False):
    return Font(name="맑은 고딕", bold=bold, size=size, color=color, italic=italic)

def aln(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ── 열 너비  B C D E F G ───────────────────
#  B=구분  C=계약기간  D=주요조건  E=비고  F=품의링크
col_widths = {1:2, 2:13, 3:20, 4:52, 5:18, 6:14, 7:2}
for col, w in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

LAST_COL = 6  # F

def sec_title(row, text):
    rh(row - 1, 7)
    rh(row, 22)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=LAST_COL)
    c = ws.cell(row=row, column=2, value=text)
    c.font      = fnt(bold=True, size=11, color=C_WHITE)
    c.fill      = fill(C_SEC_HDR)
    c.alignment = aln("left", "center")

def col_header_row(row, labels):
    rh(row, 20)
    for col, label in zip(range(2, LAST_COL + 1), labels):
        c = ws.cell(row=row, column=col, value=label)
        c.font      = fnt(bold=True, size=10, color=C_WHITE)
        c.fill      = fill(C_COL_HDR)
        c.alignment = aln("center", "center")
        c.border    = bd()

# ════════════════════════════════════════════
# 타이틀
rh(1, 8)
rh(2, 44)
ws.merge_cells("B2:F2")
c = ws.cell(row=2, column=2, value="일룸 송도점 (홈플러스 송도)  퇴점 추진 현황")
c.font      = Font(name="맑은 고딕", bold=True, size=18, color=C_WHITE)
c.fill      = fill(C_TITLE)
c.alignment = aln("center", "center")

rh(3, 21)
ws.merge_cells("B3:F3")
c = ws.cell(row=3, column=2,
    value="보고일: 2026년 6월 10일        퇴점 예정일: 2026년 8월 31일")
c.font      = Font(name="맑은 고딕", size=10, color="DDDDDD", italic=True)
c.fill      = fill(C_SUBTITLE)
c.alignment = aln("right", "center")

# ════════════════════════════════════════════
# SECTION 1: 기본 현황
sec_title(5, "① 기본 현황")

info_rows = [
    ("소재지",       "인천광역시 연수구 송도국제대로 165, 홈플러스 송도점 1층 X7·YL 코너"),
    ("개점일",       "2019년 10월 11일"),
    ("매장 유형",    "투자형 B샵  (개점 당시 S-SHOP → 현 투자형 B샵)"),
    ("계약 면적",    "444.2㎡ / 134.4평  (공유면적 포함 기준 135평)"),
    ("운영 대리점",  "이종균 대표 / (주)알루컴  (인천중앙점 겸 운영)"),
    ("계약 구조",    "전대인: 홈플러스(주)   전차인: (주)알루컴(대표전차인) + (주)일룸(공동전차인·보증인)"),
    ("현 계약 기간", "2026.03.01 ~ 2026.08.31  ▶ 계약 만료 시 갱신 없이 퇴점 예정"),
]

for i, (label, value) in enumerate(info_rows):
    r = 6 + i
    rh(r, 21)
    emph = "현 계약" in label
    bg   = C_BG_EVEN if i % 2 == 0 else C_BG_ODD

    c1 = ws.cell(row=r, column=2, value=label)
    c1.font      = fnt(bold=True)
    c1.fill      = fill(C_EMPH_BG if emph else C_LABEL)
    c1.alignment = aln("center", "center")
    c1.border    = bd()

    # B 컬럼만 값, C~F 병합
    c2 = ws.cell(row=r, column=3, value=value)
    c2.font      = Font(name="맑은 고딕", bold=emph, size=10,
                        color=C_EMPH_TXT)
    c2.fill      = fill(C_EMPH_BG if emph else bg)
    c2.alignment = aln("left", "center")
    c2.border    = bd()
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=LAST_COL)

# ════════════════════════════════════════════
# SECTION 2: 계약 이력
LINKS = {
    "최초": "https://ep.fursys.com/WebFlow/print.do?1=1&flowNo=xEGVqdLtke2AK_Y56d_fQN4YKDrsfv2GwPsXHUgIa5w",
    "연장1": "https://ep.fursys.com/WebFlow/view.do?flowNo=KC11BC_WFmrnsaftuqlULwO99NamErHHbCSyRpse9lA",
    "연장2": "https://ep.fursys.com/WebFlow/view.do?flowNo=sXOLfbZUT5VcsZNecDI7sW-M61qvppcfLGagAV8VPfY",
    "연장3": "https://ep.fursys.com/WebFlow/view.do?flowNo=gkJX6S_8GkvEqZw4k2F5v6mHWqSvl8NXoGO8jyBEkto",
    "연장4": "https://ep.fursys.com/WebFlow/view.do?flowNo=Ona4CRhGeIsauzWXP7vVabdxN08vtuJUcbyZf9XDfZE",
}

r_s2 = 14
sec_title(r_s2, "② 계약 이력  (2019 개점 → 현재)")

col_header_row(r_s2 + 1,
    ["구분", "계약 기간", "주요 조건  (VAT 별도)", "비고", "품의 링크"])

contract_history = [
    # (구분, 계약기간, 주요조건, 비고, link_key, row_h)
    ("최초\n계약",
     "2019.10.11 ~ 2020.10.10\n(12개월)",
     "수수료: 월 순매출액의 9%\n미니멈: 월매출 9,000만원 미만 시 800만원 지급\n맥시멈: 1억5천만원 초과분 수수료율 6.8% 적용",
     "S-SHOP(확장점)\n개점 품의",
     "최초", 62),
    ("",
     "2020.10.11 ~ 2022.10.10\n(24개월)",
     "수수료: 월 순매출액의 10%\n미니멈: 월매출 9,000만원 미만 시 900만원 지급\n맥시멈: 1억5천만원 초과분 수수료율 7.0% 적용",
     "조건 변경\n(8% → 9%)",
     None, 56),
    ("연장 1차",
     "2022.10.11 ~ 2023.10.31\n(약 1년)",
     "수수료: 월 순매출액의 10%\n미니멈: 월매출 9,000만원 미만 시 900만원 지급\n맥시멈: 1억5천만원 초과분 수수료율 7.0% 적용",
     "현재 운영 조건 동일",
     "연장1", 56),
    ("연장 2차",
     "2023.11.01 ~ 2024.08.31\n(약 10개월)",
     "수수료: 월 순매출액의 10%\n미니멈: 월매출 9,000만원 미만 시 900만원 지급\n맥시멈: 1억5천만원 초과분 수수료율 7.0% 적용",
     "현재 운영 조건 동일",
     "연장2", 56),
    ("연장 3차",
     "2024.09.01 ~ 2025.08.31\n(1년)",
     "수수료: 월 순매출액의 10%\n미니멈: 월매출 9,000만원 미만 시 900만원 지급\n맥시멈: 1억5천만원 초과분 수수료율 7.0% 적용",
     "직전 계약 조건 동일",
     "연장3", 56),
    ("연장 4차\n[기업회생\n대응]",
     "① 2025.09.01~2026.02.28\n② 2026.03.01~2026.08.31\n(합산 1년 재계약)",
     "① 일룸 POS / 수수료 8.2%\n   미니멈 7,380만원 (미달 시 정액) / 맥시멈 초과분 6.2%\n   임차보증금 33,325,000원 납입\n② 홈플러스 POS 복귀 / 수수료 10%\n   미니멈 9,000만원 / 맥시멈 초과분 7.0%\n   임차보증금 없음 (반환)\n※ 카드수수료 1.8% 일룸 부담 (VAT 별도)",
     "홈플러스 기업회생\n리스크 관리 목적\n★ 현행 계약",
     "연장4", 88),
    ("한시적\n조건 변경\n[상호 합의]",
     "2026.05.01 ~ 2026.06.30\n(2개월 한시 적용)",
     "수수료: 8.2% → 7.4% 로 한시 완화\n미니멈 개런티 미적용 (상호 합의) → 일괄 수수료율 7.4% 적용\n맥시멈: 1억5천만원 초과분 수수료율 5.5% 적용 (기존 6.2% → 5.5%)\n월 매출액 기준 = 결제금액 / 익월 15일 지급",
     "홈플러스 37개점\n하이퍼 휴점으로 인한\n임대 조건 완화\n갱신 없이 퇴점",
     None, 76),
]

for i, (cat, period, cond, note, link_key, h) in enumerate(contract_history):
    r = r_s2 + 2 + i
    rh(r, h)
    bg      = C_BG_EVEN if i % 2 == 0 else C_BG_ODD
    is_dark = (link_key == "연장4")       # 진회색 강조 (연장4차)
    is_amnd = (cat.startswith("한시적"))   # 한시적 변경 행 — 밝은 강조

    # 행별 색 결정
    if is_dark:
        bg_main, bg_lbl, txt = C_SEC_HDR, C_SEC_HDR, C_WHITE
    elif is_amnd:
        bg_main, bg_lbl, txt = "D6D6D6", C_EMPH_BG, C_EMPH_TXT
    else:
        bg_main, bg_lbl, txt = bg, C_LABEL, C_DARK

    # B: 구분
    c1 = ws.cell(row=r, column=2, value=cat)
    c1.font      = fnt(bold=bool(cat), size=10, color=txt)
    c1.fill      = fill(bg_lbl if not is_dark else C_SEC_HDR)
    c1.alignment = aln("center", "center")
    c1.border    = bd()

    # C: 계약 기간
    c2 = ws.cell(row=r, column=3, value=period)
    c2.font      = fnt(bold=(is_dark or is_amnd), size=10, color=txt)
    c2.fill      = fill(bg_main)
    c2.alignment = aln("center", "center")
    c2.border    = bd()

    # D: 주요 조건
    c3 = ws.cell(row=r, column=4, value=cond)
    c3.font      = fnt(bold=is_dark, size=10, color=txt)
    c3.fill      = fill(bg_main)
    c3.alignment = aln("left", "center")
    c3.border    = bd()

    # E: 비고
    c4 = ws.cell(row=r, column=5, value=note)
    c4.font      = fnt(bold=(is_dark or is_amnd), size=9, color=txt)
    c4.fill      = fill(C_SEC_HDR if is_dark else C_EMPH_BG)
    c4.alignment = aln("center", "center")
    c4.border    = bd()

    # F: 품의 링크
    c5 = ws.cell(row=r, column=6)
    if link_key and link_key in LINKS:
        c5.value     = f"품의 보기 ({link_key})"
        c5.hyperlink = LINKS[link_key]
        c5.font      = Font(name="맑은 고딕", size=9,
                            color=C_WHITE if is_dark else C_LINK,
                            underline="single", bold=is_dark)
    else:
        c5.value = "합의서"  if is_amnd else "-"
        c5.font  = fnt(size=9, color=txt if is_amnd else "999999")
    c5.fill      = fill(C_SEC_HDR if is_dark else C_BG_EVEN)
    c5.alignment = aln("center", "center")
    c5.border    = bd()

# ════════════════════════════════════════════
# SECTION 3: 홈플러스 임시휴업 이후 경과
r_s3 = r_s2 + 2 + len(contract_history) + 1
sec_title(r_s3, "③ 홈플러스 임시휴업 이후 대응 경과")

col_header_row(r_s3 + 1, ["시기", "구분", "내용", "결과", "관련 문서"])

timeline = [
    ("2026년 초",
     "홈플러스\n기업회생·임시휴업",
     "전국 37개점 임시 휴업 및 사실상 폐점 결정.\n인천 해당 점포: 가좌·송의·연수·송도·논현 등.\n마트 영업 전면 중단 / 몰 입점 임대매장은 개별 운영 지속.",
     "마트 고객 유입\n사실상 불가",
     "-"),
    ("",
     "몰 점주\n간담회 개최",
     "입점 매장 공동 대응 간담회 진행.\n▪ 임대 수수료 조정 요청\n▪ 영업시간 조정 요청  (홈플: 5/11~12 공문 안내 예정)\n▪ 몰 정상영업 온라인 게시물 업데이트 요청\n▪ 주차장 부분 폐쇄 해제 요청 (고객 유입 차단)\n▪ 고객센터 종료 → 몰 담당직원 업무 이관",
     "홈플러스 측\n공문 안내 예정",
     "-"),
    ("2026.05.14",
     "일룸→홈플러스\n공문 발송",
     "매장 운영시간 단축 협조 요청 (기존 10:00~22:00 → 단축 운영).\n사유: 야간 고객 유입 감소에 따른 운영 효율 및 인력 환경 개선.",
     "단축 운영 중",
     "발신 공문"),
    ("2026.06月",
     "원상복구\n지침서 수령",
     "홈플러스 Hyper 임시 휴업 점포 Mall 원상복구 지침서 수령.\n실질 조치: 전시품(이동 가능 집기) 회수로 완료 가능.\n별도 철거·인테리어 공사 불필요.",
     "공사 없음\n전시품 회수만",
     "지침서 수령"),
]

for i, (date, cat, content, result, doc) in enumerate(timeline):
    r = r_s3 + 2 + i
    rh(r, 70)
    bg = C_BG_EVEN if i % 2 == 0 else C_BG_ODD

    for col, val, bold, bg_c, align_h in [
        (2, date,    bool(date), bg,       "center"),
        (3, cat,     True,       C_LABEL,  "center"),
        (4, content, False,      bg,       "left"),
        (5, result,  True,       C_EMPH_BG,"center"),
        (6, doc,     False,      C_BG_EVEN,"center"),
    ]:
        c = ws.cell(row=r, column=col, value=val)
        c.font      = fnt(bold=bold, size=10)
        c.fill      = fill(bg_c)
        c.alignment = aln(align_h, "center")
        c.border    = bd()

# ════════════════════════════════════════════
# SECTION 4: 향후 계획
r_s4 = r_s3 + 2 + len(timeline) + 1
sec_title(r_s4, "④ 향후 계획")

plan_rows = [
    ("퇴점 사유",
     "홈플러스 기업회생에 따른 마트 영업 중단으로 고객 유입 사실상 불가.\n계약 만료(2026.08.31) 시 갱신 없이 철수."),
    ("퇴점 예정일",
     "2026년 8월 31일  (현 계약 만료일)"),
    ("사전 조치",
     "퇴점 공문 발송 (홈플러스 송도점 점장·몰 운영팀 수신) → 원상복구 진행"),
    ("원상복구",
     "홈플러스 지침서 기준: 전시품(집기) 회수 외 별도 공사 불필요"),
]

for i, (label, value) in enumerate(plan_rows):
    r = r_s4 + 1 + i
    rh(r, 40)
    bg   = C_BG_EVEN if i % 2 == 0 else C_BG_ODD
    emph = "퇴점 예정일" in label

    c1 = ws.cell(row=r, column=2, value=label)
    c1.font      = fnt(bold=True, color=C_WHITE if emph else C_DARK)
    c1.fill      = fill(C_SEC_HDR if emph else C_LABEL)
    c1.alignment = aln("center", "center")
    c1.border    = bd()

    c2 = ws.cell(row=r, column=3, value=value)
    c2.font      = fnt(bold=emph, size=11 if emph else 10)
    c2.fill      = fill(C_EMPH_BG if emph else bg)
    c2.alignment = aln("left", "center")
    c2.border    = bd()
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=LAST_COL)

# ════════════════════════════════════════════
# SECTION 5: 매출 흐름
r_s5 = r_s4 + len(plan_rows) + 2

sec_title(r_s5, "⑤ 매출 흐름  (2023.01 ~ 2026.06  /  단위: 백만원)")

# 구간 범례
r_leg = r_s5 + 1
rh(r_leg, 18)
ws.merge_cells(start_row=r_leg, start_column=2,
               end_row=r_leg,   end_column=LAST_COL)
c_l = ws.cell(row=r_leg, column=2,
    value="  ■ 정상 운영 (2023~2025.08)     ■ 기업회생 영향 (2025.09~2026.03)     ■ 임시휴업 급락 (2026.04~)   /   ▓ = 미니멈 미달분")
c_l.font      = fnt(size=9, italic=True, color="555555")
c_l.alignment = aln("left", "center")

# 컬럼 헤더
r_sh = r_leg + 1
rh(r_sh, 22)
for col, label in [(2,"기간"), (3,"매출\n(백만원)"),
                   (4,"매출 추이  (■ 1칸 ≈ 10백만원  /  ▓ = 미니멈 미달)"),
                   (5,"미니멈\n(백만원)"), (6,"구간 / 비고")]:
    c = ws.cell(row=r_sh, column=col, value=label)
    c.font      = fnt(bold=True, size=10, color=C_WHITE)
    c.fill      = fill(C_COL_HDR)
    c.alignment = aln("center", "center")
    c.border    = bd()

# ── 위상 설정 ──────────────────────────────
PHASE_BG     = {1:"F4F7FA", 2:"FFF8F0", 3:"FEF0EE"}
PHASE_ALT    = {1:"FFFFFF", 2:"FFFCF6", 3:"FDF6F5"}
PHASE_BARCOL = {1:"4A7EA0", 2:"D97C2A", 3:"C0392B"}
PHASE_LABEL  = {1:"정상 운영", 2:"기업회생 영향", 3:"임시휴업 이후"}

def get_phase(period):
    if period <= "2025-08": return 1
    if period <= "2026-03": return 2
    return 3

BAR_UNIT = 10   # 1블록 = 10백만원

# 데이터 (기간, 백만원, 미니멈백만, 비고)  ── 2023 추가
sales = [
    # 2023
    ("2023-01",187,90,""), ("2023-02",211,90,""), ("2023-03",105,90,""),
    ("2023-04",113,90,""), ("2023-05",107,90,""), ("2023-06", 67,90,""),
    ("2023-07", 89,90,""), ("2023-08", 94,90,""), ("2023-09", 97,90,""),
    ("2023-10",103,90,""), ("2023-11", 96,90,""), ("2023-12",167,90,""),
    # 2024
    ("2024-01",158,90,""), ("2024-02",182,90,""), ("2024-03",145,90,""),
    ("2024-04",117,90,""), ("2024-05", 99,90,""), ("2024-06",120,90,""),
    ("2024-07",137,90,""), ("2024-08",100,90,""), ("2024-09",128,90,""),
    ("2024-10", 93,90,""), ("2024-11", 84,90,""), ("2024-12",137,90,""),
    # 2025
    ("2025-01",181,90,""), ("2025-02",147,90,""), ("2025-03", 80,90,""),
    ("2025-04", 75,90,""), ("2025-05", 87,90,""), ("2025-06", 84,90,""),
    ("2025-07", 95,90,""), ("2025-08",118,90,""),
    ("2025-09", 74,74,"일룸 POS 전환"),
    ("2025-10", 83,74,""), ("2025-11", 71,74,""), ("2025-12",107,74,""),
    # 2026
    ("2026-01",121,90,""), ("2026-02",106,90,""), ("2026-03",151,90,""),
    ("2026-04", 53,90,"홈플러스 임시휴업"),
    ("2026-05", 59,90,"임시휴업 지속"),
    ("2026-06", 16,90,"부분 기간"),
]

data_start_row = r_sh + 1

for i, (period, amt, minim, note) in enumerate(sales):
    r   = data_start_row + i
    rh(r, 19)
    ph    = get_phase(period)
    bg    = PHASE_BG[ph] if i % 2 == 0 else PHASE_ALT[ph]
    under = amt < minim

    # 막대: 미니멈 이하분 ■, 초과분 ■, 미달분 ▓
    filled  = min(amt, minim) // BAR_UNIT
    over    = max(0, amt - minim) // BAR_UNIT
    deficit = (max(0, minim - amt) // BAR_UNIT) if under else 0
    bar_str = "■" * filled + ("▓" * deficit if under else "■" * over)

    # B: 기간
    c1 = ws.cell(row=r, column=2, value=period)
    c1.font = fnt(bold=(ph==3), size=10)
    c1.fill = fill(bg); c1.alignment = aln("center","center"); c1.border = bd()

    # C: 매출 (백만원)
    c2 = ws.cell(row=r, column=3, value=amt)
    c2.number_format = '0"백만"'
    c2.font = Font(name="맑은 고딕", bold=(under or ph==3), size=11,
                   color="C0392B" if under else C_EMPH_TXT)
    c2.fill = fill("FDECEA" if under else bg)
    c2.alignment = aln("center","center"); c2.border = bd()

    # D: 시각 막대
    c3 = ws.cell(row=r, column=4, value=bar_str)
    bar_color = ("C0392B" if (under and ph==3) else
                 "D97C2A" if (under and ph==2) else
                 "A02020" if under else PHASE_BARCOL[ph])
    c3.font = Font(name="맑은 고딕", size=10, color=bar_color, bold=(ph==3))
    c3.fill = fill(bg)
    c3.alignment = aln("left","center", wrap=False); c3.border = bd()

    # E: 미니멈
    c4 = ws.cell(row=r, column=5, value=minim)
    c4.number_format = '0"백만"'
    c4.font = fnt(size=9, color="888888")
    c4.fill = fill(bg); c4.alignment = aln("center","center"); c4.border = bd()

    # F: 구간/비고
    phase_note = PHASE_LABEL[ph] + (f"  ({note})" if note else "")
    c5 = ws.cell(row=r, column=6, value=phase_note)
    c5.font = Font(name="맑은 고딕", size=9, bold=(ph==3),
                   color=("C0392B" if ph==3 else "D97C2A" if ph==2 else "555555"))
    c5.fill = fill(bg); c5.alignment = aln("center","center"); c5.border = bd()

# 연도별 합계
r_sum = data_start_row + len(sales)
for yr, lbl, bg_s in [("2023","2023년 합계  (12개월)","DCDCDC"),
                       ("2024","2024년 합계  (12개월)","D4D4D4"),
                       ("2025","2025년 합계  (12개월)","CCCCCC"),
                       ("2026","2026년 합계  (1~6월)","C4C4C4")]:
    rh(r_sum, 21)
    items = [s[1] for s in sales if s[0].startswith(yr)]
    total = sum(items); avg = round(total / len(items))

    c1 = ws.cell(row=r_sum,column=2,value=lbl)
    c1.font=fnt(bold=True,size=10); c1.fill=fill(bg_s)
    c1.alignment=aln("center","center"); c1.border=bd()

    c2 = ws.cell(row=r_sum,column=3,value=total)
    c2.number_format='0"백만"'; c2.font=fnt(bold=True,size=10)
    c2.fill=fill(bg_s); c2.alignment=aln("center","center"); c2.border=bd()

    c3 = ws.cell(row=r_sum,column=4,value=f"월 평균  {avg}백만원")
    c3.font=fnt(size=9,italic=True); c3.fill=fill(bg_s)
    c3.alignment=aln("left","center"); c3.border=bd()

    for col in [5,6]:
        cx=ws.cell(row=r_sum,column=col); cx.fill=fill(bg_s); cx.border=bd()
    r_sum += 1

# ── 차트 ──────────────────────────────────────
# 미니멈 참조 숨김 컬럼 (col 8)
MCOL = 8
ws.column_dimensions[get_column_letter(MCOL)].width  = 1
ws.column_dimensions[get_column_letter(MCOL)].hidden = True
ws.cell(row=r_sh, column=MCOL, value="미니멈 기준")
for i, (_p, _a, minim, _n) in enumerate(sales):
    ws.cell(row=data_start_row + i, column=MCOL, value=minim)

from openpyxl.chart import BarChart, LineChart, Reference

# ── 연도별 평균 숨김 컬럼 (col 9) ─────────────
ACOL = 9
ws.column_dimensions[get_column_letter(ACOL)].width  = 1
ws.column_dimensions[get_column_letter(ACOL)].hidden = True
ws.cell(row=r_sh, column=ACOL, value="연도 평균")

# 연도별 평균 계산 후 월별로 채우기
yr_avg = {}
for yr in ["2023","2024","2025","2026"]:
    items = [s[1] for s in sales if s[0].startswith(yr)]
    yr_avg[yr] = round(sum(items)/len(items))
for i, (period, *_) in enumerate(sales):
    ws.cell(row=data_start_row+i, column=ACOL, value=yr_avg[period[:4]])

# ── 막대 차트 (구간별 색, 우측 배치) ───────────
bar = BarChart()
bar.type    = "col"
bar.title   = "월별 매출 추이  (2023.01 ~ 2026.06)  /  단위: 백만원"
bar.style   = 2
bar.width   = 20
bar.height  = 22
bar.y_axis.delete = True          # y축 숫자 숨김
bar.legend  = None

# x축: 모든 월 표시
bar.x_axis.tickLblSkip  = 1
bar.x_axis.tickMarkSkip = 1
bar.x_axis.numFmt       = "YY-MM"     # 23-01, 24-01 형식으로 짧게

val_ref = Reference(ws, min_col=3, min_row=r_sh,
                    max_col=3, max_row=data_start_row+len(sales)-1)
cat_ref = Reference(ws, min_col=2, min_row=data_start_row,
                    max_col=2, max_row=data_start_row+len(sales)-1)
bar.add_data(val_ref, titles_from_data=True)
bar.set_categories(cat_ref)

# 연도별 색상
YEAR_COL = {"2023":"5B9BD5",   # 파랑
            "2024":"70AD47",   # 초록
            "2025":"ED7D31",   # 주황
            "2026":"C00000"}   # 진빨강

# 막대별 색 (연도별)
for i, (period, *_) in enumerate(sales):
    yr = period[:4]
    dp = DataPoint(idx=i)
    dp.graphicalProperties.solidFill      = YEAR_COL[yr]
    dp.graphicalProperties.line.solidFill = YEAR_COL[yr]
    bar.series[0].dPt.append(dp)

# ── 연도 평균 라인 오버레이 ──────────────────
avg_line = LineChart()
a_ref = Reference(ws, min_col=ACOL, min_row=r_sh,
                  max_col=ACOL, max_row=data_start_row+len(sales)-1)
avg_line.add_data(a_ref, titles_from_data=True)
avg_line.set_categories(cat_ref)
avg_line.series[0].graphicalProperties.line.solidFill = "222222"
avg_line.series[0].graphicalProperties.line.width     = 22000   # 2.2pt
avg_line.series[0].marker.symbol = "none"

bar += avg_line   # 콤보 차트

# ── 우측 배치 ────────────────────────────────
ws.column_dimensions[get_column_letter(7)].width = 2
ws.add_chart(bar, f"H{r_s5}")

last_r = r_sum + 2

# ── 마지막 여백 ──────────────────────────────
rh(last_r, 8)

# ── 페이지 설정 ──────────────────────────────
ws.page_setup.orientation = "landscape"   # 가로로 변경 (테이블+차트 나란히)
ws.page_setup.fitToPage   = True
ws.page_setup.fitToWidth  = 1
ws.page_setup.fitToHeight = 0
ws.print_area = f"A1:{get_column_letter(14)}{last_r}"

out = r"c:\Users\FURSYS\Downloads\iloom-workspace-claude\00-inbox\송도점_퇴점추진현황_v2.xlsx"
wb.save(out)
print(f"저장 완료: {out}")
